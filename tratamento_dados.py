"""
pipeline_completo.py
====================
Executa os 2 passos em sequência:

PASSO 1 -> lê rota.xlsx, agrupa endereços -> rota_agrupada.xlsx
PASSO 2 -> consolida rotas agrupadas -> rota_processada_final.xlsx

Uso:
python pipeline_completo.py
python pipeline_completo.py --passo 2
"""

import json
import sys
import re
import unicodedata
from collections import defaultdict, OrderedDict
from difflib import SequenceMatcher
from itertools import combinations
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ARQ_ENTRADA = "rota.xlsx"
ARQ_AGRUPADO = "rota_agrupada.xlsx"
ARQ_PROCESSADO = "rota_processada_final.xlsx"

GROUP_COLORS = [
    'DCE6F1', 'E2EFDA', 'FCE4D6', 'EDD9FF', 'FFF2CC',
    'D9E1F2', 'F8CBAD', 'D6EAF8', 'FDEBD0', 'E8DAEF',
    'D5F5E3', 'FDEDEC', 'EBF5FB', 'FEF9E7', 'F4ECF7',
]

PREFIX_MAP = {
    'RUA': 'RUA', 'R': 'RUA',
    'AVENIDA': 'AVENIDA', 'AV': 'AVENIDA',
    'PRAÇA': 'PRAÇA', 'PRACA': 'PRAÇA', 'PÇ': 'PRAÇA',
    'ALAMEDA': 'ALAMEDA', 'AL': 'ALAMEDA',
    'TRAVESSA': 'TRAVESSA', 'TV': 'TRAVESSA',
}

BLDG_RE = re.compile(
    r'\b(EDI\.?|ED\.?|EF\.?|EDIF[IÍ]CIO|EDIFICIO|'
    r'CONDOM[IÍ]NIO|COND\.?|RESIDENCIAL|RESID\.?|RES\.?|BLOCO|BL\.?)\b',
    re.IGNORECASE
)

NOISE = {
    'APT', 'APTO', 'AP', 'APARTAMENTO', 'CASA', 'CS', 'LOJA', 'SALA',
    'BLOCO', 'BL', 'TORRE', 'ANDAR', 'SN', 'S/N', 'NT', 'N/T',
    'SEM', 'NUMERO', 'JD', 'JARDIM', 'DO', 'DA', 'DE', 'E', 'A', 'O',
}


def strip_accents(s: str) -> str:
    return ''.join(
        c for c in unicodedata.normalize('NFKD', s)
        if not unicodedata.combining(c)
    )


def norm_street_code(code: str) -> str:
    code = re.sub(r'([A-Z])\s*-\s*(\d)', r'\1\2', code)
    code = re.sub(r'([A-Z])\s+(\d)', r'\1\2', code)
    return code


def extract_street_key(addr: str) -> str:
    if not isinstance(addr, str):
        return ''

    u = strip_accents(addr.upper().strip())
    m = re.match(
        r'^(RUA|R\.?|AVENIDA|AV\.?|PRACA|PRAÇA|PÇ|ALAMEDA|AL\.?|TRAVESSA|TV\.?)'
        r'\s+(C[-\s]?\d+|T[-\s]?\d+|[A-Z]\s+\d+|\d+)',
        u
    )

    if not m:
        return u.split(',')[0].strip()

    raw = m.group(1).rstrip('.')
    code = norm_street_code(m.group(2).strip())
    pfx = PREFIX_MAP.get(raw, raw)
    if pfx in ('PRACA', 'PCA'):
        pfx = 'PRAÇA'
    return f"{pfx} {code}"


def _parse_qd_lt(addr: str):
    u = strip_accents(addr.upper())
    # "QD 21 LT 05" e também abreviações comuns na planilha: "Q 21 LT 05",
    # "QD21LTE05", "QD 47 LTS 01" etc — Q sozinho e LT/LTE/LTS como lote.
    c = re.search(r'\bQD?\s*(\d+[A-Z]?)\s*LT[ES]?\s*(\d+)', u)
    if c:
        return c.group(1), str(int(c.group(2)))

    qd = re.search(r'\b(?:QUADRA|QDR?\.?)\s*:?\s*(\d+[A-Z]?)', u)
    lt = re.search(r'\b(?:LOTE|LT[ES]?\.?)\s*:?\s*(\d+)', u)
    if qd and lt:
        return qd.group(1), str(int(lt.group(1)))

    return None, None


def _extract_street_number(addr: str):
    """Extrai o número de logradouro de um endereço.

    Funciona em dois modos:
    1. Ruas com código numérico no nome: "RUA 1041, 118 ..."  → captura o número
       após o prefixo+código via regex posicional.
    2. Ruas com nome por extenso: "Avenida Circular, 283, ..."  → captura o
       primeiro token numérico após a primeira vírgula, ignorando S/N e
       números de apt/sala que venham depois.
    """
    u = strip_accents(addr.upper().strip())

    # Modo 1: prefixo seguido direto de código numérico (ex: "RUA 1041, 118")
    street_re = re.match(
        r'^(?:RUA|AVENIDA|AV\.?|PRACA|PRAÇA|ALAMEDA|TRAVESSA|R\.?)'
        r'\s+(?:C[-\s]?\d+|T[-\s]?\d+|[A-Z]\s*\d+|\d+)'
        r'\s*,\s*(.*)',
        u
    )
    if street_re:
        rest = street_re.group(1)
        first_token = rest.strip().split()[0] if rest.strip() else ''
        if re.match(r'^S/?N$', first_token):
            return None
        rest_clean = re.sub(
            r'\b(?:APT|APTO|AP|APARTAMENTO|SALA|UNID|CASA)\s*\.?\s*\d+\b',
            '', rest
        )
        all_nums = re.findall(r'\b(\d+)\b', rest_clean)
        candidates = [n for n in all_nums if n not in ('0',)]
        if not candidates:
            return None
        if len(candidates[0]) >= 2:
            return candidates[0]
        longer = [n for n in candidates[1:] if len(n) >= 2]
        return longer[0] if longer else candidates[0]

    # Modo 2: rua com nome por extenso → primeiro campo após a vírgula
    parts = u.split(',')
    if len(parts) < 2:
        return None
    token_str = parts[1].strip()
    # Marcador de número por extenso antes do dígito: "Num 2", "Nº 2",
    # "N° 2", "N. 2" — tira o marcador pra sobrar só o número.
    token_str = re.sub(r'^(?:N[º°O]?\.?|NUM(?:ERO)?\.?)\s+', '', token_str)
    token = token_str.split()[0] if token_str else ''
    if re.match(r'^S/?N$', token):
        return None
    m = re.match(r'^(\d+)$', token)
    if m and len(m.group(1)) >= 2:
        return m.group(1)
    return None


def _extract_bld_name(addr: str):
    u = strip_accents(addr.upper())
    patterns = [
        r'\b(?:EDIF[IÍ]CIO|EDIFICIO)\s+([A-Z][A-Z0-9 \-\.]+?)(?:,|\s+APT|\s+APTO|\s+AP\b|\s+\d{3,}|$)',
        r'\b(?:EDI|ED|EF)\.?\s*([A-Z][A-Z0-9 \-\.]+?)(?:,|\s+APT|\s+APTO|\s+AP\b|\s+\d{3,}|$)',
        r'\b(?:RESIDENCIAL|RESID\.?)\s+([A-Z][A-Z0-9 \-\.]+?)(?:,|\s+APT|\s+APTO|\s+AP\b|\s+CASA|\s+\d{3,}|$)',
        r'\b(?:RES)\.\s*([A-Z][A-Z0-9 \-\.]+?)(?:,|\s+APT|\s+APTO|\s+AP\b|\s+\d{3,}|$)',
        r'\b(?:COND\.?|CONDOM[IÍ]NIO)\s+(?:RESIDENCIAL\s+)?([A-Z][A-Z0-9 \-\.]+?)(?:,|\s+APT|\s+APTO|\s+AP\b|\s+\d{3,}|$)',
    ]

    for pat in patterns:
        hit = re.search(pat, u)
        if hit:
            name = hit.group(1).strip().strip(',')
            name = re.sub(r'\s*-\s*$', '', name).strip()
            name = re.sub(r'\s+APT.*$', '', name).strip()
            if name and len(name) > 1:
                return name

    return None


def _bld_prefix(addr: str) -> str:
    u = addr.upper()
    if re.search(r'\b(EDIF[IÍ]CIO|EDIFICIO|EDI?\.?|EF\.?)\b', u):
        return 'ED.'
    if re.search(r'\b(RESIDENCIAL|RESID\.?|RES\.?)\b', u):
        return 'RES.'
    if re.search(r'\b(COND\.?|CONDOM[IÍ]NIO)\b', u):
        return 'COND.'
    return 'ED.'


def standardize(addr: str) -> str:
    if not isinstance(addr, str) or not addr.strip():
        return ''

    sk = extract_street_key(addr)
    qd, lt = _parse_qd_lt(addr)
    if qd is not None and lt is not None:
        return f"{sk}, {qd}-{lt}"

    num = _extract_street_number(addr)
    if num:
        return f"{sk}, {num}"

    if bool(BLDG_RE.search(addr)):
        name = _extract_bld_name(addr)
        if name:
            return f"{sk}, {_bld_prefix(addr)} {name}"

    return sk


def _standardized_key(addr: str) -> str:
    return standardize(addr)


def fuzzy_key(addr: str) -> str:
    if not isinstance(addr, str):
        return ''

    t = strip_accents(addr.upper())
    t = re.sub(r'^(RUA|AVENIDA|AV\.?|PRACA|PRAÇA|PÇ|ALAMEDA|TRAVESSA|R\.?)\s+', '', t.strip())
    t = norm_street_code(t)
    t = BLDG_RE.sub('', t)
    t = re.sub(r'\b(APT|APTO|AP|APARTAMENTO|CASA|CS)\s*\d+\b', '', t)
    t = re.sub(r'\b\d+\b', '', t)
    t = re.sub(r'[,\.\-/\\|]', ' ', t)
    tokens = [tk for tk in t.split() if tk and tk not in NOISE and len(tk) > 1]
    return ' '.join(sorted(tokens))


def similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()


def group_rows_p1(rows, threshold=0.72):
    n = len(rows)
    parent = list(range(n))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(x, y):
        rx, ry = find(x), find(y)
        if rx != ry:
            parent[rx] = ry

    def get_num(k):
        m = re.search(r',\s*(\d+)', k)
        return m.group(1) if m else None

    streets = [extract_street_key(r['address']) for r in rows]
    std_keys = [_standardized_key(r['address']) for r in rows]
    fkeys = [fuzzy_key(r['address']) for r in rows]
    bldnames = [strip_accents(_extract_bld_name(r['address']) or '') for r in rows]

    for i, j in combinations(range(n), 2):
        if streets[i] != streets[j]:
            continue

        ki, kj = std_keys[i], std_keys[j]
        ki_has_detail = ',' in ki
        kj_has_detail = ',' in kj

        if ki_has_detail and kj_has_detail:
            if ki == kj:
                union(i, j)
            continue

        num_i = get_num(ki)
        num_j = get_num(kj)

        if num_i is not None and num_j is not None and num_i != num_j:
            continue

        if (num_i is not None and not kj_has_detail) or (num_j is not None and not ki_has_detail):
            continue

        # Guarda extra: compara número direto do endereço bruto,
        # evita agrupar "Rua X, 290" com "Rua X, 546" quando std_key falhou em extrair o número
        raw_num_i = _extract_street_number(rows[i]['address'])
        raw_num_j = _extract_street_number(rows[j]['address'])
        if raw_num_i is not None and raw_num_j is not None and raw_num_i != raw_num_j:
            continue

        same_bld = bldnames[i] and bldnames[j] and bldnames[i] == bldnames[j]

        # Quando nenhum dos dois lados tem QUALQUER detalhe extraído (nem
        # número, nem qd/lt, nem nome de prédio reconhecido) e a rua é a
        # mesma, o único critério que sobra é o texto livre — e aí um
        # limite de 0.72 é frouxo demais: "Rua X, S/N, prédio A" e
        # "Rua X, Q21 Lt05" (quando a extração falha) acabam se parecendo
        # só por compartilharem o nome da rua. Nesse caso exige uma
        # parecença bem mais alta (praticamente o mesmo texto).
        sem_detalhe_algum = not ki_has_detail and not kj_has_detail and not same_bld
        limiar = 0.93 if sem_detalhe_algum else threshold

        if similarity(fkeys[i], fkeys[j]) >= limiar or same_bld:
            union(i, j)

    groups = defaultdict(list)
    for i in range(n):
        groups[find(i)].append(i)

    for root, members in groups.items():
        labels = [standardize(rows[m]['address']) for m in members]
        best = max(labels, key=lambda x: (len(x), x.count(','), x))
        stops = sorted(
            [rows[m]['stop'] for m in members if rows[m]['stop']],
            key=lambda s: str(s).lstrip('-').zfill(10)
        )
        stop_str = ', '.join(str(s) for s in stops)

        for m in members:
            rows[m]['group_id'] = root
            rows[m]['group_label'] = best
            rows[m]['group_stops'] = f"Stop: {stop_str}" if len(stops) > 1 else ''
            rows[m]['group_size'] = len(members)

    return rows


def load_excel_p1(path: str):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value or '').strip() for c in ws[1]]

    def find_col(pats):
        for pat in pats:
            for i, h in enumerate(headers):
                if re.search(pat, h, re.IGNORECASE):
                    return i
        return None

    col_stop = find_col([r'stop', r'^c$', r'^parada'])
    col_addr = find_col([r'destination.?address', r'endere[cç]o', r'address'])
    col_lat = find_col([r'\blat\b', r'latitude'])
    col_lon = find_col([r'\blon\b', r'\blng\b', r'longitude'])

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        def g(idx):
            if idx is None or idx >= len(row):
                return ''
            v = row[idx]
            return str(v).strip() if v is not None else ''

        lat, lon = g(col_lat), g(col_lon)
        coord = f"{lat},{lon}" if lat and lon else ''

        rows.append({
            'raw_row': list(row),
            'stop': g(col_stop),
            'address': g(col_addr),
            'coord': coord,
        })

    return rows, headers, min(len(headers), 12)


def write_excel_p1(rows, headers, orig_cols, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rotas Agrupadas"

    thin = Side(style='thin', color='BFBFBF')
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
    orig_hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    orig_fill = PatternFill('solid', start_color='1F4E79')

    current_col = 1
    for i in range(orig_cols):
        if i == 4:
            continue
        cell = ws.cell(row=1, column=current_col, value=headers[i])
        cell.font = orig_hdr_font
        cell.fill = orig_fill
        cell.alignment = CTR
        cell.border = bdr
        current_col += 1

    new_hdrs = [
        (13, 'ENDERECO_REFORMADO', '1F6B39'),
        (14, 'ENDERECO_ORIGINAL', '1F4E79'),
        (15, 'ROTAS_IGUAIS', '7B3F00'),
        (16, 'STOPs DO GRUPO', '4A235A'),
        (17, 'OBSERVACAO', '1A5276'),
        (18, 'COORDENADAS', '404040'),
    ]

    for col, title, color in new_hdrs:
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = orig_hdr_font
        cell.fill = PatternFill('solid', start_color=color)
        cell.alignment = CTR
        cell.border = bdr

    ws.row_dimensions[1].height = 30
    unique_gids = list(dict.fromkeys(r.get('group_id', i) for i, r in enumerate(rows)))
    gid_color = {gid: GROUP_COLORS[i % len(GROUP_COLORS)] for i, gid in enumerate(unique_gids)}

    for i, row in enumerate(rows, 2):
        gid = row.get('group_id', i)
        color = gid_color.get(gid, 'FFFFFF')
        rfill = PatternFill('solid', start_color=color)
        grouped = row.get('group_size', 1) > 1

        current_col = 1
        for c_idx in range(orig_cols):
            if c_idx == 4:
                continue
            val = row['raw_row'][c_idx] if c_idx < len(row['raw_row']) else ''
            cell = ws.cell(row=i, column=current_col, value=val)
            cell.font = Font(name='Arial', size=9)
            cell.fill = rfill
            cell.border = bdr
            cell.alignment = CTR if current_col == 1 else LFT
            current_col += 1

        c = ws.cell(row=i, column=13, value=row.get('group_label', ''))
        c.font = Font(name='Arial', size=10, bold=True)
        c.fill = rfill
        c.border = bdr
        c.alignment = LFT

        c = ws.cell(row=i, column=14, value=row['raw_row'][4] if len(row['raw_row']) > 4 else '')
        c.font = Font(name='Arial', size=9)
        c.fill = rfill
        c.border = bdr
        c.alignment = LFT

        c = ws.cell(row=i, column=15, value=row.get('group_size', 1))
        c.font = Font(name='Arial', size=10, bold=grouped)
        c.border = bdr
        c.alignment = CTR
        c.fill = PatternFill('solid', start_color='FFC000') if grouped else rfill

        c = ws.cell(row=i, column=16, value=row.get('group_stops', ''))
        c.font = Font(name='Arial', size=9, color='4A235A', bold=grouped)
        c.fill = rfill
        c.border = bdr
        c.alignment = LFT

        c = ws.cell(row=i, column=17, value='AGRUPADO' if grouped else '')
        c.font = Font(name='Arial', size=9, bold=grouped)
        c.fill = rfill
        c.border = bdr
        c.alignment = CTR

        c = ws.cell(row=i, column=18, value=row.get('coord', ''))
        c.font = Font(name='Arial', size=9)
        c.fill = rfill
        c.border = bdr
        c.alignment = CTR

    widths = {
        'A': 18, 'B': 8, 'C': 8, 'D': 20, 'E': 18, 'F': 14, 'G': 14, 'H': 14, 'I': 14,
        'J': 14, 'K': 14, 'L': 14, 'M': 44, 'N': 50, 'O': 13, 'P': 34, 'Q': 12, 'R': 26
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:R{ws.max_row}"
    wb.save(out_path)


def passo1():
    import os
    if not os.path.exists(ARQ_ENTRADA):
        print(f"[ERRO] Arquivo nao encontrado: {ARQ_ENTRADA}")
        return False

    print(f"\n{'='*60}")
    print(" PASSO 1 - Agrupamento de enderecos")
    print(f" Entrada : {ARQ_ENTRADA}")
    print(f" Saida   : {ARQ_AGRUPADO}")
    print(f"{'='*60}")
    print(f"Lendo: {ARQ_ENTRADA}")

    rows, headers, orig_cols = load_excel_p1(ARQ_ENTRADA)

    print("Processando...")
    rows = group_rows_p1(rows)

    print(f"Salvando: {ARQ_AGRUPADO}")
    write_excel_p1(rows, headers, orig_cols, ARQ_AGRUPADO)
    print("Passo 1 concluido!\n")
    return True


def _round_coord(val: str, decimals: int = 4) -> str:
    try:
        return str(round(float(val.strip()), decimals))
    except ValueError:
        return val.strip()


def dedup_coords(coord_list: list) -> list:
    seen = {}
    for raw in coord_list:
        raw = raw.strip()
        if not raw:
            continue
        parts = raw.split(',')
        if len(parts) != 2:
            seen.setdefault((raw,), raw)
            continue
        key = (_round_coord(parts[0]), _round_coord(parts[1]))
        seen.setdefault(key, raw)
    return list(seen.values())


def dedup_single(val_list: list) -> list:
    seen = {}
    for v in val_list:
        v = v.strip()
        if not v:
            continue
        seen.setdefault(_round_coord(v), v)
    return list(seen.values())


def load_saida_p2(path: str):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value or '').strip() for c in ws[1]]

    def col_idx(pats):
        for pat in pats:
            for i, h in enumerate(headers):
                if re.search(pat, h, re.IGNORECASE):
                    return i
        return None

    idx_reformado = col_idx([r'ENDERE.O_REFORMADO', r'REFORMADO'])
    idx_original = col_idx([r'ENDERE.O_ORIGINAL', r'ORIGINAL'])
    idx_bairro = col_idx([r'bairro', r'district', r'neighborhood'])
    idx_seq = col_idx([r'sequence', r'sequen', r'\bseq\b'])
    idx_stop = col_idx([r'stop'])
    idx_coord = col_idx([r'coordenadas', r'coord'])
    idx_zip = col_idx([r'zip', r'postal', r'cep'])
    idx_lat = col_idx([r'\blat\b', r'latitude'])
    idx_lon = col_idx([r'\blon\b', r'\blng\b', r'longitude'])
    # Telefone do destinatário (vem do CSV do Anjun como coluna "Contact",
    # preservada intacta pelo Passo 1). Não existe coluna de nome no CSV
    # de origem — só o número mesmo.
    idx_contato = col_idx([r'contact', r'contato', r'telefone', r'\bphone\b'])

    if idx_reformado is None:
        raise ValueError("Coluna ENDERECO_REFORMADO nao encontrada no arquivo.")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        def g(i):
            if i is None or i >= len(row):
                return ''
            v = row[i]
            return str(v).strip() if v is not None else ''

        rows.append({
            'reformado': g(idx_reformado),
            'original': g(idx_original),
            'bairro': g(idx_bairro),
            'seq': g(idx_seq),
            'stop': g(idx_stop),
            'coord': g(idx_coord),
            'zip': g(idx_zip),
            'lat': g(idx_lat),
            'lon': g(idx_lon),
            'contato': g(idx_contato),
        })

    return rows, headers


def consolidate_p2(rows):
    groups = OrderedDict()
    for row in rows:
        base_key = row['reformado'].strip().upper() if row['reformado'] else '__SEM_ENDERECO__'
        bairro_norm = (row['bairro'] or '').strip().upper()
        # Endereços com o mesmo texto (rua/número) mas em bairros diferentes
        # são endereços DIFERENTES — ex.: "RUA C160, 312-18" pode existir em
        # dois setores distintos da cidade. O bairro entra na chave de
        # agrupamento pra não misturar esses casos no mesmo grupo.
        group_key = (base_key, bairro_norm) if base_key != '__SEM_ENDERECO__' else (base_key, '')
        groups.setdefault(group_key, []).append(row)

    result = []
    for (base_key, _bairro_norm), members in groups.items():
        stops = sorted(
            [(m['seq'] or m['stop']) for m in members if (m['seq'] or m['stop'])],
            key=lambda s: str(s).lstrip('-').zfill(10)
        )

        seqs = sorted(
            [m['seq'] for m in members if m['seq']],
            key=lambda s: str(s).lstrip('-').zfill(10)
        )

        seq_val = ', '.join(seqs)
        all_coords = [m['coord'] for m in members if m['coord']]
        unique_coords = dedup_coords(all_coords)
        lat_vals = dedup_single([m['lat'] for m in members if m['lat']])
        lon_vals = dedup_single([m['lon'] for m in members if m['lon']])
        zip_val = next((m['zip'] for m in members if m['zip']), '')

        originals_seen = []
        for m in members:
            o = m['original'].strip()
            if o and o not in originals_seen:
                originals_seen.append(o)
        original_val = ' | '.join(originals_seen)

        bairro_val = next((m['bairro'] for m in members if m['bairro']), '')

        # Endereço final (o que vira "Destination Address" e é usado pelo
        # app/geocodificação): acrescenta o bairro no fim, pra desambiguar
        # números repetidos em setores diferentes.
        # Ex.: "RUA C160, 312-18" -> "RUA C160, 312-18, JARDIM AMÉRICA"
        key_display = base_key if base_key != '__SEM_ENDERECO__' else ''
        bairro_fmt = bairro_val.strip().upper()
        if key_display and bairro_fmt and bairro_fmt not in key_display:
            key_display = f"{key_display}, {bairro_fmt}"

        # Lista detalhada de cada membro do grupo (seq + endereço original
        # + contato/telefone), usada pelo app para permitir desagrupar um
        # item específico e exibir o telefone de cada entrega junto com
        # seu endereço. Usa 'seq' (coluna SEQUENCE) pois é o valor único
        # por linha; 'stop' pode vir de uma coluna auxiliar com valores
        # repetidos.
        membros = [
            {
                'stop': m['seq'] or m['stop'],
                'original': m['original'] or m['reformado'],
                'contato': m['contato'],
            }
            for m in members
            if m['seq'] or m['stop'] or m['original'] or m['reformado']
        ]

        # Telefones únicos do grupo, na ordem em que aparecem — quando o
        # grupo tem uma única parada isso vira "o" telefone da entrega;
        # quando tem várias, cada uma mantém o seu próprio em 'membros'
        # e aqui fica a lista consolidada (útil pra exibir no resumo do
        # grupo sem abrir a lista de membros).
        contatos_seen = []
        for m in members:
            c = (m['contato'] or '').strip()
            if c and c not in contatos_seen:
                contatos_seen.append(c)
        contato_val = ', '.join(contatos_seen)

        result.append({
            'seq': seq_val,
            'key': key_display,
            'original': original_val,
            'bairro': bairro_val,
            'stops': stops,
            'coords': unique_coords,
            'zip': zip_val,
            'lat': lat_vals[0] if lat_vals else '',
            'lon': lon_vals[0] if lon_vals else '',
            'coord': unique_coords[0] if unique_coords else '',
            'count': len(members),
            'membros': membros,
            'contato': contato_val,
        })

    return result


def write_excel_p2(groups, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rotas Consolidadas"

    thin = Side(style='thin', color='BFBFBF')
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
    hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)

    col_defs = [
        ('SEQUENCE', '2E4057', 10),
        ('Destination Address', '1F6B39', 44),
        ('ROTAS_IGUAIS', '7B3F00', 13),
        ('STOPs DO GRUPO', '4A235A', 38),
        ('ZIP / POSTAL', '6B4226', 14),
        ('LATITUDE', '1A5276', 18),
        ('LONGITUDE', '117A65', 18),
        ('COORDENADAS', '404040', 28),
        ('OBSERVACAO', '6C3483', 14),
        ('ENDERECO_ORIGINAL', '8B0000', 44),
        ('BAIRRO', '2E5984', 25),
        ('CONTATO', '145A32', 18),
        ('MEMBROS_JSON', 'FFFFFF', 10),
    ]

    for ci, (title, color, width) in enumerate(col_defs, 1):
        c = ws.cell(row=1, column=ci, value=title)
        c.font = hdr_font
        c.fill = PatternFill('solid', start_color=color)
        c.alignment = CTR
        c.border = bdr
        ws.column_dimensions[get_column_letter(ci)].width = width
        if title == 'MEMBROS_JSON':
            ws.column_dimensions[get_column_letter(ci)].hidden = True

    ws.row_dimensions[1].height = 28

    for ri, grp in enumerate(groups, 2):
        count = grp['count']
        grouped = count > 1
        color = GROUP_COLORS[(ri - 2) % len(GROUP_COLORS)]
        rfill = PatternFill('solid', start_color=color)
        ofill = PatternFill('solid', start_color='FFC000')
        stop_str = ', '.join(grp['stops']) if grp['stops'] else ''
        obs = 'AGRUPADO' if grouped else ''

        cells = [
            (1, grp['seq'], CTR, Font(name='Arial', size=9), rfill),
            (2, grp['key'], LFT, Font(name='Arial', size=10, bold=grouped), rfill),
            (3, count, CTR, Font(name='Arial', size=10, bold=grouped), ofill if grouped else rfill),
            (4, stop_str, LFT, Font(name='Arial', size=9, color='4A235A', bold=grouped), rfill),
            (5, grp['zip'], CTR, Font(name='Arial', size=9), rfill),
            (6, grp['lat'], CTR, Font(name='Arial', size=9), rfill),
            (7, grp['lon'], CTR, Font(name='Arial', size=9), rfill),
            (8, grp['coord'], CTR, Font(name='Arial', size=9), rfill),
            (9, obs, CTR, Font(name='Arial', size=9, bold=grouped), rfill),
            (10, grp['original'], LFT, Font(name='Arial', size=9), rfill),
            (11, grp['bairro'], LFT, Font(name='Arial', size=9), rfill),
            (12, grp.get('contato', ''), CTR, Font(name='Arial', size=9), rfill),
            (13, json.dumps(grp.get('membros', []), ensure_ascii=False), LFT, Font(name='Arial', size=8), rfill),
        ]

        for ci, val, aln, fnt, fill in cells:
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = fnt
            c.fill = fill
            c.alignment = aln
            c.border = bdr

        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(col_defs))}{ws.max_row}"
    wb.save(out_path)
    print(f"Salvo: {out_path} ({len(groups)} enderecos unicos)")


def passo2():
    import os
    if not os.path.exists(ARQ_AGRUPADO):
        print(f"[ERRO] Arquivo nao encontrado: {ARQ_AGRUPADO}")
        return False

    print(f"\n{'='*60}")
    print(" PASSO 2 - Consolidacao de rotas")
    print(f" Entrada : {ARQ_AGRUPADO}")
    print(f" Saida   : {ARQ_PROCESSADO}")
    print(f"{'='*60}")
    print(f"Lendo: {ARQ_AGRUPADO}")

    rows, headers = load_saida_p2(ARQ_AGRUPADO)
    print(f" {len(rows)} linhas lidas")

    groups = consolidate_p2(rows)
    print(f" {len(groups)} enderecos unicos | {sum(1 for g in groups if g['count'] > 1)} com duplicatas")

    print(f"Salvando: {ARQ_PROCESSADO}")
    write_excel_p2(groups, ARQ_PROCESSADO)
    print("Passo 2 concluido!\n")
    return True


def main():
    passos = [1, 2]

    for arg in sys.argv[1:]:
        if arg.startswith('--passo'):
            val = arg.split('=')[-1] if '=' in arg else sys.argv[sys.argv.index(arg) + 1]
            passos = [int(p.strip()) for p in val.split(',')]

    print("\n" + "=" * 60)
    print(" PIPELINE DE PROCESSAMENTO DE ROTAS")
    print(f" Passos a executar: {passos}")
    print("=" * 60)

    if 1 in passos:
        ok = passo1()
        if not ok and 2 in passos:
            print("[ABORTADO] Passo 1 falhou.")
            return

    if 2 in passos:
        passo2()

    print("\n" + "=" * 60)
    print(" PIPELINE FINALIZADO")
    print("=" * 60)


if __name__ == '__main__':
    main()