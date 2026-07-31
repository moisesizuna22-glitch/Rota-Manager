"""
servidor.py  —  Rota Manager  (FastAPI)
========================================
Substitui o http.server original por FastAPI + Uvicorn.
Todas as rotas, lógica de negócio, planos e integrações
foram preservadas 100%.  O frontend (rota_manager1.html)
não precisa de nenhuma alteração.

Rodando localmente:
    pip install fastapi uvicorn python-multipart requests openpyxl
    python servidor.py

Railway / Render / Fly:
    A plataforma define PORT via variável de ambiente.
    Configure também: HERE_API_KEY, BREVO_API_KEY, BREVO_SENDER_EMAIL,
    DATA_DIR, ADMIN_PASS, ORS_API_KEY, GOOGLE_GEOCODING_API_KEY
    (opcionais mas recomendados em produção).

    ORS_API_KEY: chave gratuita do OpenRouteService (openrouteservice.org),
    usada em /rota/otimizar. Sem ela, cai automaticamente para o servidor
    público de demonstração do OSRM, que não tem SLA e pode falhar.
"""

import base64
import csv
import hashlib
import json
import os
import re
import secrets
import shutil
import subprocess
import sys
import threading
import time
import uuid
import zipfile
from datetime import datetime, timedelta
from pathlib import Path

import requests
import uvicorn
from fastapi import FastAPI, Request, Header, HTTPException, UploadFile, File
from fastapi.responses import JSONResponse, FileResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles

import gamification
from csv_para_rota_xlsx import _extrair_cep, _extrair_numero, _geocodificar_enderecos

# ═══════════════════════════════════════════════════════════════════
#  CONFIGURAÇÃO
# ═══════════════════════════════════════════════════════════════════

HOST           = os.environ.get("HOST", "0.0.0.0")
PORT           = int(os.environ.get("PORT", 8792))
HTML_FILE      = "rota_manager1.html"
ARQ_ENTRADA    = "rota.xlsx"
ARQ_PROCESSADO = "rota_processada_final.xlsx"
ARQ_VALIDADO   = "rota_validada_here.xlsx"
TRATAMENTO_PY  = "tratamento_dados.py"
ANJUN_SCRIPT   = "csv_para_rota_xlsx.py"
ANJUN_TRATAMENTO_PY = "anjun_tratamento.py"
ANJUN_EXTRATOR_PY = "anjun_extrator.py"
IMILE_EXTRATOR_PY = "imile_extrator.py"
IMILE_TRATAMENTO_PY = "imile_tratamento.py"

HERE_API_KEY   = os.environ.get("HERE_API_KEY", "P8C0izk0pJ1PIZr3d5CpeAI8b_dc7YFLkNKJlzP0A-M").strip()
HERE_CIDADE_UF = os.environ.get("HERE_CIDADE_UF", "Goiânia - GO, Brasil").strip()

# Chave do Google Maps JS/Places pro navegador (mapa visual + autosuggest) —
# diferente da GOOGLE_GEOCODING_API_KEY abaixo, que é server-side/geocoding
# só. Cai pro valor que já estava hardcoded no frontend se a env var não
# estiver configurada, pra não quebrar o deploy atual.
GOOGLE_MAPS_BROWSER_KEY = os.environ.get("GOOGLE_MAPS_BROWSER_KEY", "AIzaSyAaESp78MrYmySlHiupJKjjgCi0-1dfWnE").strip()

# Token público do Mapbox (prefixo "pk." já indica que é pra rodar no
# navegador — diferente das chaves acima). Ainda assim só é liberado via
# /api/config, igual as outras, pra manter um único caminho de distribuição.
MAPBOX_BROWSER_TOKEN = os.environ.get("MAPBOX_BROWSER_TOKEN", "pk.eyJ1IjoibW9pc2Vzc2VuanVpem8iLCJhIjoiY21xOHRzOXZyMDFidTJyb2x0Zml5a2FxNyJ9.EHf8W94vDiim9_EDVGlP1g").strip()

OSRM_BASE_URL  = os.environ.get("OSRM_BASE_URL", "https://router.project-osrm.org")

ORS_API_KEY    = os.environ.get("ORS_API_KEY", "").strip()
ORS_BASE_URL   = "https://api.openrouteservice.org"

GOOGLE_VISION_API_KEY = (os.environ.get("GOOGLE_VISION_API_KEY") or "").strip() or None
GOOGLE_VISION_URL     = "https://vision.googleapis.com/v1/images:annotate"

# Usada só como fallback de geocodificação (depois que o HERE falha), e só
# pra endereços "simples" (rua + número), pra economizar a cota. Configure
# GOOGLE_GEOCODING_API_KEY no Railway pra ativar; sem ela, o app segue só
# com HERE + banco de coordenadas normalmente.
GOOGLE_GEOCODING_API_KEY = os.environ.get("GOOGLE_GEOCODING_API_KEY", "").strip()
GOOGLE_GEOCODING_URL     = "https://maps.googleapis.com/maps/api/geocode/json"

BREVO_API_URL       = "https://api.brevo.com/v3/smtp/email"
BREVO_API_KEY       = os.environ.get("BREVO_API_KEY")
BREVO_SENDER_EMAIL  = os.environ.get("BREVO_SENDER_EMAIL")
BREVO_SENDER_NOME   = os.environ.get("BREVO_SENDER_NOME", "Rota Manager")
EMAIL_TTL_MINUTOS   = 10

DATA_DIR = Path(os.environ.get("DATA_DIR", "/data" if Path("/data").is_dir() else "."))
DATA_DIR.mkdir(parents=True, exist_ok=True)

USERS_FILE        = str(DATA_DIR / "usuarios.json")
HISTORICO_FILE    = str(DATA_DIR / "historico_rotas.json")
BANCO_COORDS_FILE = str(DATA_DIR / "banco_coords.json")

# ─── Backup ──────────────────────────────────────────────────────────
# Pasta de snapshots DENTRO do mesmo volume (protege contra escrita
# corrompida / sobrescrita acidental), + envio periódico por email
# (BREVO) para fora do Railway (protege contra perda do volume inteiro).
BACKUP_DIR               = DATA_DIR / "backups"
BACKUP_DIR.mkdir(parents=True, exist_ok=True)
BACKUP_MAX_SNAPSHOTS     = int(os.environ.get("BACKUP_MAX_SNAPSHOTS", "40"))  # por arquivo
BACKUP_EMAIL_DESTINO     = os.environ.get("BACKUP_EMAIL_DESTINO", "").strip()  # ex: seu@email.com
BACKUP_EMAIL_INTERVALO_H = float(os.environ.get("BACKUP_EMAIL_INTERVALO_H", "24"))

PAGAMENTO_LINK_REUSE_MINUTOS  = 30
INFINITEPAY_HANDLE            = "moisessenju"
INFINITEPAY_LINKS_URL         = "https://api.checkout.infinitepay.io/links"
INFINITEPAY_PAYMENT_CHECK_URL = "https://api.checkout.infinitepay.io/payment_check"

# ─── Planos ────────────────────────────────────────────────────────
PLANOS = {
    "avulsa": {
        "nome":      "Importação Avulsa",
        "preco":     2.00,
        "tipo":      "avulso",
        "dias":      None,
        "beneficio": "Use 1 importação avulsa",
        "badge":     None,
        "pagamento_automatico": True,
        "importacoes_por_dia": None,
    },
    "essencial": {
        "nome":      "Plano Essencial",
        "preco":     30.00,
        "tipo":      "mensal",
        "dias":      30,
        "beneficio": "1 importação por dia",
        "badge":     None,
        "pagamento_automatico": True,
        "importacoes_por_dia": 1,
    },
    "profissional": {
        "nome":      "Plano Profissional",
        "preco":     60.00,
        "tipo":      "mensal",
        "dias":      30,
        "beneficio": "2 importações por dia",
        "badge":     "MAIS POPULAR",
        "pagamento_automatico": True,
        "importacoes_por_dia": 2,
    },
}

SESSION_TTL_HORAS = 12
DIAS_TESTE_GRATIS = 3
AVISO_TRIAL_HORAS_ANTES = 24  # avisa quando faltar 1 dia (ou menos) pro teste grátis acabar

def _conceder_teste_gratis(user: dict) -> None:
    """3 dias grátis automáticos pra TODA conta nova, não importa por onde
    ela foi criada (cadastro público com email, ou criação direta pelo
    painel admin) — acesso completo (sem plano_ativo, então sem limite
    diário). "origem_acesso" só serve pro painel admin diferenciar teste
    grátis de plano pago; some sozinho quando o acesso expirar ou quando
    a pessoa assinar/for liberada de verdade (ver admin_liberar_acesso e
    _creditar_plano)."""
    expira_trial = datetime.now() + timedelta(days=DIAS_TESTE_GRATIS)
    user["acesso_expira_em"] = expira_trial.isoformat()
    user["origem_acesso"]    = "teste_gratis"

# ═══════════════════════════════════════════════════════════════════
#  FASTAPI APP
# ═══════════════════════════════════════════════════════════════════

app = FastAPI(title="Rota Manager", docs_url=None, redoc_url=None)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["GET", "POST", "DELETE", "OPTIONS"],
    allow_headers=["*"],
)

# ─── Gamificação ────────────────────────────────────────────────────
# (o include_router() fica lá embaixo, DEPOIS da rota /api/perfil/me —
# senão a rota genérica /api/perfil/{user_id} do gamification.router
# "rouba" a palavra "me" como se fosse um user_id)

_ICONS_DIR = Path(__file__).parent / "static" / "icons"
if _ICONS_DIR.is_dir():
    app.mount("/static/icons", StaticFiles(directory=str(_ICONS_DIR)), name="icons")


def _garantir_perfil(user_id: str) -> "gamification.PerfilUsuario":
    """Cria o perfil de gamificação na primeira vez que o usuário aparece."""
    perfil = gamification.carregar_usuario(user_id)
    if perfil is None:
        perfil = gamification.PerfilUsuario(user_id=user_id)
        gamification.salvar_usuario(perfil)
    return perfil


def _conceder_xp_rota(user_id: str, paradas: int, pacotes: int, rota_hash: str) -> dict | None:
    """Concede XP por rota concluída. Nunca deixa um erro de gamificação
    derrubar o pipeline principal — só loga e segue."""
    try:
        _garantir_perfil(user_id)
        payload = gamification.RegistrarRotaPayload(
            user_id=user_id, paradas_concluidas=paradas,
            pacotes_entregues=pacotes, rota_hash=rota_hash,
        )
        resultado = gamification.registrar_conclusao_rota(payload)
        if resultado.xp_ganho:
            print(f"  [XP] {user_id} +{resultado.xp_ganho} XP (rota) "
                  f"— nível {resultado.perfil.nivel}"
                  f"{' 🎉 subiu de nível!' if resultado.subiu_de_nivel else ''}")
        return resultado.model_dump()
    except Exception as e:
        print(f"  [XP] ⚠️ falha ao conceder XP de rota pra {user_id}: {e}")
        return None


def _conceder_xp_endereco(user_id: str, endereco: str) -> dict | None:
    try:
        _garantir_perfil(user_id)
        payload = gamification.RegistrarEnderecoPayload(user_id=user_id, enderecos=[endereco])
        resultado = gamification.registrar_endereco_corrigido(payload)
        if resultado.xp_ganho:
            print(f"  [XP] {user_id} +{resultado.xp_ganho} XP (endereço) "
                  f"— nível {resultado.perfil.nivel}"
                  f"{' 🎉 subiu de nível!' if resultado.subiu_de_nivel else ''}")
        return resultado.model_dump()
    except Exception as e:
        print(f"  [XP] ⚠️ falha ao conceder XP de endereço pra {user_id}: {e}")
        return None

def ok_json(data: dict, status: int = 200) -> JSONResponse:
    return JSONResponse(content=data, status_code=status)

def err_json(msg: str, status: int = 400) -> JSONResponse:
    return JSONResponse(content={"ok": False, "erro": msg}, status_code=status)

def _base_url(request: Request) -> str:
    host = request.headers.get("x-forwarded-host") or request.headers.get("host") or f"localhost:{PORT}"
    if host.split(":")[0] in ("localhost", "127.0.0.1"):
        proto = "http"
    else:
        proto = request.headers.get("x-forwarded-proto", "https")
    return f"{proto}://{host}"

# ═══════════════════════════════════════════════════════════════════
#  BACKUP  (snapshot local rotativo + cópia fora do volume por email)
# ═══════════════════════════════════════════════════════════════════

def _backup_snapshot(path: Path):
    """Antes de sobrescrever um arquivo de dados, guarda uma cópia com
    timestamp em DATA_DIR/backups/. Mantém só os últimos
    BACKUP_MAX_SNAPSHOTS por arquivo (apaga os mais antigos)."""
    try:
        if not path.exists():
            return
        ts = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
        destino = BACKUP_DIR / f"{path.stem}_{ts}{path.suffix}"
        shutil.copy2(path, destino)
        # poda: mantém só os N mais recentes desse arquivo
        irmaos = sorted(
            BACKUP_DIR.glob(f"{path.stem}_*{path.suffix}"),
            key=lambda p: p.stat().st_mtime,
        )
        for antigo in irmaos[:-BACKUP_MAX_SNAPSHOTS]:
            antigo.unlink(missing_ok=True)
    except Exception as e:
        print(f"  [backup] falha ao criar snapshot de {path.name}: {e}")

def _snapshot_mais_recente(path: Path) -> Path | None:
    candidatos = sorted(
        BACKUP_DIR.glob(f"{path.stem}_*{path.suffix}"),
        key=lambda p: p.stat().st_mtime,
    )
    return candidatos[-1] if candidatos else None

def _salvar_com_backup(path_str: str, conteudo_serializado: str, dados_novos_vazios: bool, forcar: bool = False):
    """Escrita segura e comum a usuarios/historico/banco_coords:
    1) tira snapshot do estado atual antes de mexer (mesmo com forcar=True
       — só o BLOQUEIO abaixo é pulado, o snapshot pra recuperação manual
       continua sendo tirado sempre);
    2) recusa sobrescrever um arquivo COM dados por um payload vazio
       (a causa mais comum de perda de dados: um bug ou uma corrida
       de leitura fizeram o app achar que estava tudo vazio) — a menos
       que forcar=True, pra ações de limpeza deliberadas e confirmadas
       pelo admin (ex.: limpar histórico de todo mundo de propósito)."""
    path = Path(path_str)
    if path.exists() and dados_novos_vazios and not forcar:
        try:
            tinha_conteudo = len(path.read_text("utf-8").strip()) > 2  # mais que "{}"/"[]"
        except Exception:
            tinha_conteudo = True  # arquivo ilegível: por segurança, trata como "tinha dado"
        if tinha_conteudo:
            print(f"  [backup] BLOQUEADO: recusando sobrescrever {path.name} (tinha dados) "
                  f"com payload vazio. Nada foi alterado — investigue antes de forçar.")
            return False
    _backup_snapshot(path)
    path.write_text(conteudo_serializado, "utf-8")
    return True

def _carregar_com_recuperacao(path_str: str, vazio_padrao):
    """Lê um arquivo de dados; se ele não existir ou estiver corrompido,
    tenta se recuperar sozinho a partir do snapshot de backup mais
    recente (em vez de silenciosamente devolver vazio, que é o que
    causava a perda em cascata: 'não achei nada' → próxima escrita
    grava vazio 'de verdade')."""
    path = Path(path_str)
    if path.exists():
        try:
            return json.loads(path.read_text("utf-8"))
        except Exception as e:
            print(f"  [backup] {path.name} corrompido ({e}); tentando restaurar de snapshot...")
    else:
        print(f"  [backup] {path.name} não existe; tentando restaurar de snapshot (se houver)...")
    recente = _snapshot_mais_recente(path)
    if recente:
        try:
            dados = json.loads(recente.read_text("utf-8"))
            print(f"  [backup] restaurado {path.name} a partir de {recente.name}")
            # regrava o arquivo principal com o que foi recuperado
            path.write_text(recente.read_text("utf-8"), "utf-8")
            return dados
        except Exception as e:
            print(f"  [backup] snapshot {recente.name} também corrompido: {e}")
    return vazio_padrao

def _zip_backup_atual() -> Path:
    """Empacota os 3 arquivos de dados atuais num .zip em /tmp, pra
    anexar no email de backup."""
    zip_path = Path("/tmp") / f"backup_rota_manager_{datetime.now().strftime('%Y-%m-%dT%H-%M-%S')}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in (USERS_FILE, HISTORICO_FILE, BANCO_COORDS_FILE):
            p = Path(f)
            if p.exists():
                zf.write(p, arcname=p.name)
    return zip_path

def enviar_backup_por_email(destino: str = "") -> tuple[bool, str]:
    """Manda os arquivos de dados atuais, zipados, por email via Brevo —
    cópia fora do volume do Railway. Configure BACKUP_EMAIL_DESTINO
    (ou passe `destino`) e BREVO_API_KEY/BREVO_SENDER_EMAIL."""
    destino = (destino or BACKUP_EMAIL_DESTINO).strip()
    if not destino:
        return False, "BACKUP_EMAIL_DESTINO não configurado."
    if not BREVO_API_KEY or not BREVO_SENDER_EMAIL:
        return False, "BREVO_API_KEY/BREVO_SENDER_EMAIL ausentes."
    try:
        zip_path = _zip_backup_atual()
        conteudo_b64 = base64.b64encode(zip_path.read_bytes()).decode("ascii")
        payload = {
            "sender":      {"name": BREVO_SENDER_NOME, "email": BREVO_SENDER_EMAIL},
            "to":          [{"email": destino}],
            "subject":     f"[Rota Manager] Backup automático — {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            "textContent": "Backup automático dos dados (usuarios.json, historico_rotas.json, banco_coords.json) em anexo.",
            "attachment":  [{"content": conteudo_b64, "name": zip_path.name}],
        }
        headers = {"api-key": BREVO_API_KEY, "Content-Type": "application/json", "Accept": "application/json"}
        r = requests.post(BREVO_API_URL, json=payload, headers=headers, timeout=30)
        zip_path.unlink(missing_ok=True)
        if r.status_code in (200, 201):
            return True, "Backup enviado."
        return False, f"Brevo retornou erro {r.status_code}: {r.text[:200]}"
    except Exception as e:
        return False, f"Falha ao enviar backup: {e}"

def _loop_backup_periodico():
    """Thread em background: manda backup por email a cada
    BACKUP_EMAIL_INTERVALO_H horas. Só roda se BACKUP_EMAIL_DESTINO
    estiver configurado."""
    if not BACKUP_EMAIL_DESTINO:
        print("  [backup] BACKUP_EMAIL_DESTINO não configurado — backup por email desativado "
              "(defina essa variável de ambiente no Railway pra ativar).")
        return
    intervalo_s = max(BACKUP_EMAIL_INTERVALO_H, 1) * 3600
    while True:
        ok, msg = enviar_backup_por_email()
        print(f"  [backup] envio periódico: {'ok' if ok else 'falhou'} — {msg}")
        time.sleep(intervalo_s)

# ═══════════════════════════════════════════════════════════════════
#  MIGRAÇÃO DE VOLUME
# ═══════════════════════════════════════════════════════════════════

def _migrar_para_volume():
    if str(DATA_DIR) == ".":
        return
    for nome in ("usuarios.json", "historico_rotas.json", "banco_coords.json"):
        destino = DATA_DIR / nome
        origem  = Path(nome)
        if not destino.exists() and origem.exists():
            try:
                destino.write_text(origem.read_text("utf-8"), "utf-8")
                print(f"  [migração] {nome} → volume ({DATA_DIR})")
            except Exception as e:
                print(f"  [migração] falha {nome}: {e}")

# ═══════════════════════════════════════════════════════════════════
#  SESSÕES EM MEMÓRIA
# ═══════════════════════════════════════════════════════════════════

_sessoes: dict = {}

def _limpar_sessoes_expiradas():
    agora = datetime.now()
    expiradas = [t for t, s in _sessoes.items()
                 if agora - s["criado_em"] > timedelta(hours=SESSION_TTL_HORAS)]
    for t in expiradas:
        del _sessoes[t]

def criar_sessao(user_id: str, usuario: str, is_admin: bool = False) -> str:
    _limpar_sessoes_expiradas()
    token = secrets.token_hex(32)
    _sessoes[token] = {
        "user_id":   user_id,
        "usuario":   usuario,
        "is_admin":  is_admin,
        "dados":     None,
        "criado_em": datetime.now(),
    }
    return token

def obter_sessao(token: str) -> dict | None:
    _limpar_sessoes_expiradas()
    s = _sessoes.get(token)
    if s is None:
        return None
    if datetime.now() - s["criado_em"] > timedelta(hours=SESSION_TTL_HORAS):
        del _sessoes[token]
        return None
    return s

def destruir_sessao(token: str):
    _sessoes.pop(token, None)

def _token_da_request(request: Request) -> str | None:
    auth = request.headers.get("authorization", "")
    if auth.startswith("Bearer "):
        return auth[7:].strip()
    return None

def _sessao_ou_401(request: Request) -> dict:
    token = _token_da_request(request)
    if not token:
        raise HTTPException(status_code=401, detail="Não autenticado.")
    sess = obter_sessao(token)
    if sess is None:
        raise HTTPException(status_code=401, detail="Sessão expirada ou inválida. Faça login novamente.")
    return sess

def _sessao_admin_ou_403(request: Request) -> dict:
    sess = _sessao_ou_401(request)
    if not sess.get("is_admin"):
        raise HTTPException(status_code=403, detail="Acesso restrito ao administrador.")
    return sess

def _sessao_com_acesso_ou_403(request: Request) -> dict:
    sess = _sessao_ou_401(request)
    if sess.get("is_admin"):
        return sess
    if not usuario_tem_acesso_ativo(sess["usuario"]):
        raise HTTPException(status_code=403,
            detail="Sua assinatura venceu. Assine para continuar importando rotas.")
    pode, motivo = usuario_pode_importar_hoje(sess["usuario"])
    if not pode:
        raise HTTPException(status_code=403, detail=motivo)
    return sess

# ─── Converte HTTPException em JSON padrão do app ──────────────────
@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    return JSONResponse(
        status_code=exc.status_code,
        content={"ok": False, "erro": exc.detail},
    )

# ═══════════════════════════════════════════════════════════════════
#  EMAIL  (Brevo)
# ═══════════════════════════════════════════════════════════════════

def _email_valido(email: str) -> bool:
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email or ""))

def _gerar_codigo() -> str:
    return f"{secrets.randbelow(1_000_000):06d}"

def enviar_codigo_email(destino: str, codigo: str) -> tuple[bool, str]:
    if not BREVO_API_KEY or not BREVO_SENDER_EMAIL:
        return False, "Servidor não configurado para enviar email (BREVO_API_KEY/BREVO_SENDER_EMAIL ausentes)."
    try:
        payload = {
            "sender":      {"name": BREVO_SENDER_NOME, "email": BREVO_SENDER_EMAIL},
            "to":          [{"email": destino}],
            "subject":     "Seu código de verificação — Rota Manager",
            "textContent": (
                f"Seu código de verificação do Rota Manager é: {codigo}\n\n"
                f"Esse código expira em {EMAIL_TTL_MINUTOS} minutos.\n"
                f"Se você não solicitou este cadastro, ignore este email."
            ),
        }
        headers = {"api-key": BREVO_API_KEY, "Content-Type": "application/json", "Accept": "application/json"}
        r = requests.post(BREVO_API_URL, json=payload, headers=headers, timeout=15)
        if r.status_code in (200, 201):
            return True, ""
        return False, f"Brevo retornou erro {r.status_code}: {r.text[:200]}"
    except Exception as e:
        return False, f"Falha ao enviar email: {e}"

# ═══════════════════════════════════════════════════════════════════
#  CADASTRO PENDENTE (verificação por email)
# ═══════════════════════════════════════════════════════════════════

_cadastros_pendentes: dict = {}

def _limpar_cadastros_expirados():
    agora = datetime.now()
    expirados = [t for t, c in _cadastros_pendentes.items()
                 if agora - c["criado_em"] > timedelta(minutes=EMAIL_TTL_MINUTOS)]
    for t in expirados:
        del _cadastros_pendentes[t]

def _telefone_valido(tel: str) -> bool:
    digits = re.sub(r"[\s\-().]+", "", tel or "")
    return bool(re.match(r"^\d{2}9\d{8}$", digits))

def _normalizar_telefone(tel: str) -> str:
    return re.sub(r"[\s\-().]+", "", tel or "")

def iniciar_cadastro_pendente(username: str, email: str, senha: str, telefone: str = "") -> tuple[bool, str, str | None]:
    _limpar_cadastros_expirados()
    username = username.strip()
    email    = email.strip().lower()
    telefone = _normalizar_telefone(telefone) if telefone else ""

    if not username or len(username) < 3:
        return False, "Usuário deve ter pelo menos 3 caracteres.", None
    if not _email_valido(email):
        return False, "Email inválido.", None
    if not senha or len(senha) < 4:
        return False, "Senha deve ter pelo menos 4 caracteres.", None
    if telefone and not _telefone_valido(telefone):
        return False, "Telefone inválido. Use DDD + 9 + número (ex: 62 9 91153473).", None

    users = carregar_usuarios()
    chave_existente, _ = _buscar_usuario(users, username)
    if chave_existente is not None:
        return False, "Usuário já existe.", None
    if any(u.get("email", "").lower() == email for u in users.values()):
        return False, "Este email já está cadastrado em outra conta.", None

    codigo = _gerar_codigo()
    ok, erro = enviar_codigo_email(email, codigo)
    if not ok:
        return False, erro, None

    pending_token = secrets.token_hex(16)
    _cadastros_pendentes[pending_token] = {
        "username":   username,
        "email":      email,
        "telefone":   telefone,
        "senha_hash": _hash_senha(senha),
        "codigo":     codigo,
        "tentativas": 0,
        "criado_em":  datetime.now(),
    }
    return True, "Código enviado para o email.", pending_token

def confirmar_cadastro(pending_token: str, codigo: str) -> tuple[bool, str]:
    _limpar_cadastros_expirados()
    pend = _cadastros_pendentes.get(pending_token)
    if pend is None:
        return False, "Cadastro expirado ou inválido. Solicite um novo código."
    pend["tentativas"] += 1
    if pend["tentativas"] > 5:
        del _cadastros_pendentes[pending_token]
        return False, "Muitas tentativas incorretas. Solicite um novo código."
    if codigo.strip() != pend["codigo"]:
        return False, "Código incorreto."
    users = carregar_usuarios()
    chave_existente, _ = _buscar_usuario(users, pend["username"])
    if chave_existente is not None:
        del _cadastros_pendentes[pending_token]
        return False, "Usuário já existe."
    users[pend["username"]] = {
        "id":       str(uuid.uuid4()),
        "hash":     pend["senha_hash"],
        "email":    pend["email"],
        "telefone": pend.get("telefone", ""),
    }
    # 3 dias grátis automáticos pra toda conta nova (ver _conceder_teste_gratis).
    _conceder_teste_gratis(users[pend["username"]])
    salvar_usuarios(users)
    del _cadastros_pendentes[pending_token]
    return True, f"Conta criada com sucesso. Você ganhou {DIAS_TESTE_GRATIS} dias grátis!"

# ═══════════════════════════════════════════════════════════════════
#  RECUPERAÇÃO DE SENHA
# ═══════════════════════════════════════════════════════════════════

_recuperacoes_pendentes: dict = {}

def _limpar_recuperacoes_expiradas():
    agora = datetime.now()
    expirados = [t for t, c in _recuperacoes_pendentes.items()
                 if agora - c["criado_em"] > timedelta(minutes=EMAIL_TTL_MINUTOS)]
    for t in expirados:
        del _recuperacoes_pendentes[t]

def _buscar_usuario_por_login_ou_email(users: dict, identificador: str):
    alvo = (identificador or "").strip().lower()
    if not alvo:
        return None, None
    chave, dados = _buscar_usuario(users, identificador)
    if dados is not None:
        return chave, dados
    for chave, dados in users.items():
        if dados.get("email", "").lower() == alvo:
            return chave, dados
    return None, None

def iniciar_recuperacao_senha(identificador: str) -> tuple[bool, str, str | None]:
    _limpar_recuperacoes_expiradas()
    users = carregar_usuarios()
    chave, u = _buscar_usuario_por_login_ou_email(users, identificador)
    if u is None:
        return False, "Usuário ou email não encontrado.", None
    email = u.get("email", "").strip()
    if not email:
        return False, ("Esta conta não possui email cadastrado para recuperação automática. "
                       "Peça ao administrador para redefinir sua senha."), None
    codigo = _gerar_codigo()
    ok, erro = enviar_codigo_email(email, codigo)
    if not ok:
        return False, erro, None
    recovery_token = secrets.token_hex(16)
    _recuperacoes_pendentes[recovery_token] = {
        "username":   chave,
        "email":      email,
        "codigo":     codigo,
        "tentativas": 0,
        "criado_em":  datetime.now(),
    }
    em_user, _, em_dom = email.partition("@")
    mascarado = (em_user[:2] + "***@" + em_dom) if len(em_user) > 2 else ("***@" + em_dom)
    return True, mascarado, recovery_token

def confirmar_codigo_recuperacao(recovery_token: str, codigo: str) -> tuple[bool, str]:
    _limpar_recuperacoes_expiradas()
    pend = _recuperacoes_pendentes.get(recovery_token)
    if pend is None:
        return False, "Solicitação expirada ou inválida. Comece novamente."
    pend["tentativas"] += 1
    if pend["tentativas"] > 5:
        del _recuperacoes_pendentes[recovery_token]
        return False, "Muitas tentativas incorretas. Solicite um novo código."
    if codigo.strip() != pend["codigo"]:
        return False, "Código incorreto."
    pend["confirmado"] = True
    return True, "Código confirmado."

def redefinir_senha_recuperacao(recovery_token: str, nova_senha: str) -> tuple[bool, str]:
    _limpar_recuperacoes_expiradas()
    pend = _recuperacoes_pendentes.get(recovery_token)
    if pend is None:
        return False, "Solicitação expirada ou inválida. Comece novamente."
    if not pend.get("confirmado"):
        return False, "Confirme o código antes de definir a nova senha."
    if not nova_senha or len(nova_senha) < 4:
        return False, "A nova senha deve ter pelo menos 4 caracteres."
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, pend["username"])
    if u is None:
        del _recuperacoes_pendentes[recovery_token]
        return False, "Usuário não encontrado."
    u["hash"] = _hash_senha(nova_senha)
    salvar_usuarios(users)
    del _recuperacoes_pendentes[recovery_token]
    return True, "Senha redefinida com sucesso. Faça login com a nova senha."

# ═══════════════════════════════════════════════════════════════════
#  HISTÓRICO
# ═══════════════════════════════════════════════════════════════════

def carregar_historico() -> list:
    return _carregar_com_recuperacao(HISTORICO_FILE, [])

def salvar_historico(historico: list, forcar: bool = False):
    _salvar_com_backup(
        HISTORICO_FILE,
        json.dumps(historico, ensure_ascii=False, indent=2),
        dados_novos_vazios=(len(historico) == 0),
        forcar=forcar,
    )

def adicionar_ao_historico(nome_arquivo: str, rows: list, headers: list, user_id: str = ""):
    historico = carregar_historico()
    existia_antes = any(h.get("nome") == nome_arquivo and h.get("user_id") == user_id for h in historico)
    entrada = {
        "nome":     nome_arquivo,
        "total":    len(rows),
        "headers":  headers,
        "rows":     rows,
        "user_id":  user_id,
        "salvo_em": datetime.now().strftime("%d/%m/%Y %H:%M"),
    }
    historico = [h for h in historico if not (h.get("nome") == nome_arquivo and h.get("user_id") == user_id)]
    historico.insert(0, entrada)
    historico = historico[:50]
    salvar_historico(historico)
    if existia_antes:
        # A rota anterior desse usuário saiu do histórico (foi substituída
        # por essa nova importação) — as correções manuais dele eram só
        # pra ela, então voltam ao valor global normal.
        banco_coords_limpar_overrides_usuario(user_id)
    return entrada

def atualizar_rows_historico(nome_arquivo: str, rows: list, user_id: str = "") -> bool:
    """Atualiza (in-place) as rows de uma entrada JÁ existente no histórico,
    sem disparar a limpeza de overrides — usado só pra refletir, depois do
    /pipeline, os endereços que a confirmação de geolocalização (HERE/
    Google) achou. Não conta como "a rota saiu do histórico"."""
    historico = carregar_historico()
    for h in historico:
        if h.get("nome") == nome_arquivo and h.get("user_id") == user_id:
            h["rows"]  = rows
            h["total"] = len(rows)
            salvar_historico(historico)
            return True
    return False

# ═══════════════════════════════════════════════════════════════════
#  BANCO DE COORDENADAS
# ═══════════════════════════════════════════════════════════════════

def _normalizar_endereco(end: str) -> str:
    return re.sub(r"\s+", " ", (end or "").strip().lower())

def _dist_metros(lat1, lon1, lat2, lon2):
    """Distância aproximada em metros entre dois pontos (haversine) — mesma
    fórmula usada na busca de quadra/lote, só que como helper reaproveitável
    aqui pro consenso do banco de coordenadas."""
    import math
    R = 6371000
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlmb = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlmb / 2) ** 2
    return R * 2 * math.atan2(a ** 0.5, (1 - a) ** 0.5)

# Tolerância pra considerar duas coordenadas "o mesmo lugar" (nível de
# precisão de endereço/prédio, não de quadra inteira).
BANCO_COORDS_TOLERANCIA_M = 40

def banco_coords_carregar() -> dict:
    """
    Formato do arquivo:
        {
          "global":      { chave: {lat, lon, endereco_original, salvo_em, fonte,
                                    usuario, confirmacoes, pendente?} },
          "overrides":   { user_id: { chave: {lat, lon, endereco_original, salvo_em} } },
          "correlacoes": { chave: {grupo_id, endereco_original, quando, por} }
        }
    "global" é o banco compartilhado por todo mundo — alimentado tanto por
    confirmação de API (HERE/Google) quanto por correção manual (arrastar o
    pin), o "ciclo" que vai pegando os prédios de Goiânia aos poucos.
    "confirmacoes" conta quantas contas distintas já bateram no valor atual;
    "pendente" (opcional) guarda um valor DIFERENTE que outro usuário propôs
    mas que ainda não foi corroborado por uma segunda conta — só substitui o
    valor ativo quando alguém mais confirma o mesmo ponto (dentro da
    tolerância de BANCO_COORDS_TOLERANCIA_M), pra 1 correção errada não
    estragar um valor que outras pessoas já estão usando. "overrides" é a
    view PESSOAL de cada usuário (o que ele vê na hora, mesmo antes de
    qualquer promoção pro global) — vale só enquanto a rota que contém
    aquele endereço estiver no histórico dele; quando ela sai do histórico
    (apagada ou substituída por nova importação), o override é descartado e
    ele volta a ver o valor global normal. "correlacoes" é a marcação manual
    de "esses textos são o mesmo lugar" (usuário agrupou 2+ paradas no mapa
    ou na lista) — é só uma DICA de consenso, nunca sobrescreve coordenada
    de ninguém sozinha (ver banco_coords_correlacionar).
    """
    banco = _carregar_com_recuperacao(
        BANCO_COORDS_FILE, {"global": {}, "overrides": {}, "correlacoes": {}}
    )
    # Migração do formato antigo (dict plano na raiz, sem separar override).
    if "global" not in banco and "overrides" not in banco:
        banco = {"global": banco, "overrides": {}}
    banco.setdefault("global", {})
    banco.setdefault("overrides", {})
    banco.setdefault("correlacoes", {})
    return banco

def banco_coords_salvar(banco: dict):
    vazio = not banco.get("global") and not banco.get("overrides") and not banco.get("correlacoes")
    _salvar_com_backup(
        BANCO_COORDS_FILE,
        json.dumps(banco, ensure_ascii=False, indent=2),
        dados_novos_vazios=vazio,
    )

def banco_coords_promover_manual(banco: dict, chave: str, endereco_original: str, lat: float, lon: float, user_id: str) -> str:
    """Tenta promover uma correção manual (pin arrastado) pro banco global
    compartilhado. Modifica banco["global"] in-place. Retorna uma palavra
    descrevendo o que aconteceu, só pra compor a mensagem de retorno:

    - "novo": não havia valor global pra esse endereço — vira o valor global
      direto, ninguém mais estava confiando em outra coisa.
    - "proprio": o valor global atual já era desse mesmo usuário — atualiza
      direto (ele tá corrigindo o próprio ajuste, não pisando no de outro).
    - "confirmado": o novo ponto bate (dentro da tolerância) com o valor
      global já existente — reforça a confiança, sem precisar mudar nada.
    - "promovido": o novo ponto diverge do valor global atual, MAS uma
      segunda conta diferente já tinha proposto esse mesmo ponto antes
      (ficou "pendente") — agora com 2 contas concordando, vira o valor
      ativo pra todo mundo.
    - "pendente_novo": o novo ponto diverge do valor global atual e é o
      primeiro (ou outro) a discordar — fica marcado como "pendente" nesse
      endereço, mas NÃO troca o valor que os outros já estão usando; só
      troca se uma segunda conta confirmar esse mesmo ponto depois.

    Também atualiza banco["stats_usuario"][user_id] = {"confirmadas", "corrigidas"}
    — só nos dois casos com veredito resolvido ("confirmado" e "promovido"),
    já que "novo"/"proprio"/"pendente_novo" ainda não foram julgados por
    ninguém. É a base do stat de precisão que o admin vê no painel (ver
    admin_listar_usuarios) — não é público pros outros usuários, só o
    admin, pra não transformar "errar um endereço" em constrangimento
    público e desincentivar gente de tentar corrigir.
    """
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    globais = banco["global"]
    atual = globais.get(chave)
    stats = banco.setdefault("stats_usuario", {})

    def _bump(uid: str, campo: str):
        if not uid:
            return
        s = stats.setdefault(uid, {"confirmadas": 0, "corrigidas": 0})
        s[campo] = s.get(campo, 0) + 1

    if atual is None:
        globais[chave] = {
            "endereco_original": endereco_original.strip(),
            "lat": round(lat, 6), "lon": round(lon, 6),
            "salvo_em": agora, "fonte": "manual",
            "usuario": user_id, "confirmacoes": 1,
        }
        return "novo"

    mesmo_dono = atual.get("usuario") == user_id and atual.get("fonte") == "manual"
    bate = _dist_metros(atual["lat"], atual["lon"], lat, lon) <= BANCO_COORDS_TOLERANCIA_M

    if mesmo_dono or bate:
        atual["lat"] = round(lat, 6)
        atual["lon"] = round(lon, 6)
        atual["salvo_em"] = agora
        atual["fonte"] = atual.get("fonte") or "manual"
        if bate and not mesmo_dono:
            atual["confirmacoes"] = atual.get("confirmacoes", 1) + 1
            _bump(atual.get("usuario"), "confirmadas")  # dono original foi corroborado
        atual.pop("pendente", None)
        return "confirmado" if (bate and not mesmo_dono) else "proprio"

    # Diverge de um valor que outra conta/fonte já confia — não troca sozinho.
    pendente = atual.get("pendente")
    if pendente and pendente.get("usuario") != user_id and \
       _dist_metros(pendente["lat"], pendente["lon"], lat, lon) <= BANCO_COORDS_TOLERANCIA_M:
        dono_antigo = atual.get("usuario")
        _bump(dono_antigo, "corrigidas")          # valor antigo foi derrubado por 2 contas
        _bump(pendente.get("usuario"), "confirmadas")  # quem propôs o valor certo
        _bump(user_id, "confirmadas")                  # quem confirmou por último, mesma coisa
        globais[chave] = {
            "endereco_original": endereco_original.strip(),
            "lat": round(lat, 6), "lon": round(lon, 6),
            "salvo_em": agora, "fonte": "manual",
            "usuario": user_id, "confirmacoes": 2,
        }
        return "promovido"

    atual["pendente"] = {"lat": round(lat, 6), "lon": round(lon, 6), "usuario": user_id, "quando": agora}
    return "pendente_novo"

def banco_coords_salvar_coord(endereco: str, lat: float, lon: float, user_id: str) -> tuple[bool, str, dict]:
    """Correção manual (usuário arrastou o pin no mapa). Sempre atualiza a
    view pessoal do usuário na hora (vale só enquanto a rota que contém esse
    endereço estiver no histórico dele — veja banco_coords_limpar_overrides_usuario);
    também tenta promover pro banco global compartilhado, ver
    banco_coords_promover_manual pras regras de quando isso é imediato ou
    fica pendente até uma segunda pessoa confirmar."""
    chave = _normalizar_endereco(endereco)
    if not chave:
        return False, "Endereço vazio.", {}
    try:
        lat = float(lat)
        lon = float(lon)
    except (TypeError, ValueError):
        return False, "Coordenadas inválidas.", {}
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    banco = banco_coords_carregar()
    user_id = user_id or "(desconhecido)"

    overrides_usuario = banco["overrides"].setdefault(user_id, {})
    entrada = overrides_usuario.get(chave) or {"endereco_original": endereco.strip()}
    entrada["lat"] = round(lat, 6)
    entrada["lon"] = round(lon, 6)
    entrada["endereco_original"] = entrada.get("endereco_original") or endereco.strip()
    entrada["salvo_em"] = agora
    overrides_usuario[chave] = entrada

    resultado = banco_coords_promover_manual(banco, chave, endereco, lat, lon, user_id)

    banco_coords_salvar(banco)
    print(f"  [BANCO_COORDS] override de {user_id!r} em {chave!r} → ({lat:.6f}, {lon:.6f}) [{resultado}]")

    mensagens = {
        "novo":          "Coordenada salva — já entrou pro banco compartilhado.",
        "proprio":       "Coordenada atualizada — já valia pra todo mundo, agora com o valor novo.",
        "confirmado":    "Coordenada salva e confirmada — reforça o valor que já era compartilhado.",
        "promovido":     "Coordenada confirmada por uma segunda pessoa — agora vale pra todo mundo.",
        "pendente_novo": "Coordenada salva (já vale pra você) — precisa outra pessoa confirmar o mesmo ponto antes de virar o valor compartilhado de todos.",
    }
    msg = mensagens.get(resultado, "Coordenada salva.")
    return True, msg, {"lat": entrada["lat"], "lon": entrada["lon"]}

def banco_coords_salvar_global(endereco: str, lat: float, lon: float, fonte: str, usuario: str = "") -> dict:
    """Coordenada confirmada por API (HERE ou Google) — fica permanente no
    banco compartilhado. Só muda de novo se um usuário der override nela."""
    chave = _normalizar_endereco(endereco)
    if not chave:
        return {}
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    banco = banco_coords_carregar()
    entrada = banco["global"].get(chave) or {"endereco_original": endereco.strip()}
    entrada["lat"] = round(float(lat), 6)
    entrada["lon"] = round(float(lon), 6)
    entrada["endereco_original"] = entrada.get("endereco_original") or endereco.strip()
    entrada["salvo_em"] = agora
    entrada["fonte"]    = fonte
    if usuario:
        entrada["usuario"] = usuario
    banco["global"][chave] = entrada
    banco_coords_salvar(banco)
    print(f"  [BANCO_COORDS] confirmado via {fonte} → {chave!r} = ({entrada['lat']:.6f}, {entrada['lon']:.6f})")
    return {"lat": entrada["lat"], "lon": entrada["lon"]}

def banco_coords_buscar(endereco: str, user_id: str = "") -> dict | None:
    """Prioriza o override pessoal do usuário (se a rota ainda estiver no
    histórico dele); senão usa o valor global confirmado por API. Retorna
    None se não achar nada — aí sim vale a pena chamar HERE/Google."""
    chave = _normalizar_endereco(endereco)
    if not chave:
        return None
    banco = banco_coords_carregar()
    if user_id:
        override = banco["overrides"].get(user_id, {}).get(chave)
        if override:
            return {"lat": override["lat"], "lon": override["lon"], "fonte": "override"}
    entrada = banco["global"].get(chave)
    if entrada:
        return {"lat": entrada["lat"], "lon": entrada["lon"], "fonte": entrada.get("fonte", "banco")}
    return None

def banco_coords_limpar_overrides_usuario(user_id: str):
    """Chamado quando a rota que tinha as correções manuais sai do
    histórico do usuário (foi apagada ou substituída por uma nova
    importação) — as correções eram só pra aquela rota, então o usuário
    volta a ver o valor global normal."""
    if not user_id:
        return
    banco = banco_coords_carregar()
    if banco["overrides"].pop(user_id, None) is not None:
        banco_coords_salvar(banco)
        print(f"  [BANCO_COORDS] overrides de {user_id!r} limpos (rota saiu do histórico)")

def banco_coords_apagar(endereco: str) -> tuple[bool, str]:
    """Remove do banco global — usado só pelo painel admin."""
    chave = _normalizar_endereco(endereco)
    banco = banco_coords_carregar()
    if chave not in banco["global"]:
        return False, "Endereço não encontrado no banco."
    del banco["global"][chave]
    banco_coords_salvar(banco)
    return True, "Entrada removida do banco."

def banco_coords_apagar_recentes(horas: float) -> int:
    """Remove do banco global toda entrada confirmada nas últimas N horas —
    usado quando uma leva de geocodificações saiu errada (ex.: HERE
    confundiu quadra/lote) e o Geni quer forçar tudo a ser reconfirmado."""
    limite = datetime.now() - timedelta(hours=horas)
    banco = banco_coords_carregar()
    removidas = []
    for chave, entrada in banco["global"].items():
        salvo_em = entrada.get("salvo_em", "")
        try:
            dt = datetime.strptime(salvo_em, "%d/%m/%Y %H:%M")
        except (ValueError, TypeError):
            continue  # sem data confiável, não mexe
        if dt >= limite:
            removidas.append(chave)
    for chave in removidas:
        del banco["global"][chave]
    if removidas:
        banco_coords_salvar(banco)
    return len(removidas)

def banco_coords_apagar_intervalo(data_inicio: str, data_fim: str) -> int:
    """
    Remove do banco global toda entrada cujo "salvo_em" caiu dentro do
    intervalo [data_inicio 00:00, data_fim 23:59:59] (inclusive nas duas
    pontas). data_inicio/data_fim vêm como "YYYY-MM-DD" (formato de
    <input type="date">). Ex.: apagar tudo confirmado de 28/06/2026 até
    hoje, pra forçar reconfirmação de uma leva que saiu errada.
    """
    try:
        inicio = datetime.strptime(data_inicio, "%Y-%m-%d")
        fim = datetime.strptime(data_fim, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except (ValueError, TypeError):
        raise ValueError("Datas inválidas — use o formato YYYY-MM-DD.")
    if fim < inicio:
        raise ValueError("Data final não pode ser antes da data inicial.")
    banco = banco_coords_carregar()
    removidas = []
    for chave, entrada in banco["global"].items():
        salvo_em = entrada.get("salvo_em", "")
        try:
            dt = datetime.strptime(salvo_em, "%d/%m/%Y %H:%M")
        except (ValueError, TypeError):
            continue  # sem data confiável, não mexe
        if inicio <= dt <= fim:
            removidas.append(chave)
    for chave in removidas:
        del banco["global"][chave]
    if removidas:
        banco_coords_salvar(banco)
    return len(removidas)

def banco_coords_correlacionar(enderecos: list[str], user_id: str = "") -> dict:
    """
    Correlação manual entre endereços (usuário agrupou 2+ paradas — no
    mapa ou na lista — que na verdade são o mesmo lugar, só escritas
    diferente). É só uma DICA pra consenso futuro: nunca sobrescreve
    coordenada de ninguém sozinha; só serve pra avisar quando um desses
    endereços aparecer de novo sem coordenada própria, mas outro do
    mesmo grupo já tiver (ver o uso em banco_coords_aplicar).

    Cada endereço só pode estar em 1 grupo. Se algum dos endereços
    informados já pertencia a um grupo (de uma correlação anterior),
    todo mundo se junta a esse grupo — inclusive fundindo 2 grupos
    antigos diferentes em 1 só, se for o caso.
    """
    pares: list[tuple[str, str]] = []
    vistas = set()
    for e in enderecos or []:
        chave = _normalizar_endereco(e)
        if chave and chave not in vistas:
            vistas.add(chave)
            pares.append((chave, (e or "").strip()))
    if len(pares) < 2:
        return {"ok": False, "msg": "Precisa de pelo menos 2 endereços diferentes pra correlacionar."}

    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    banco = banco_coords_carregar()
    correlacoes = banco["correlacoes"]

    grupos_existentes = {correlacoes[c]["grupo_id"] for c, _ in pares if c in correlacoes}
    if grupos_existentes:
        grupo_id = sorted(grupos_existentes)[0]
        if len(grupos_existentes) > 1:
            # fundindo grupos antigos diferentes em 1 só (o vencedor é o "menor" id)
            for info in correlacoes.values():
                if info.get("grupo_id") in grupos_existentes:
                    info["grupo_id"] = grupo_id
    else:
        grupo_id = uuid.uuid4().hex[:12]

    for chave, original in pares:
        entrada = correlacoes.get(chave) or {}
        entrada["grupo_id"]          = grupo_id
        entrada["endereco_original"] = entrada.get("endereco_original") or original
        entrada["quando"]            = agora
        if user_id:
            entrada["por"] = user_id
        correlacoes[chave] = entrada

    banco_coords_salvar(banco)
    membros = sorted({
        info["endereco_original"] for info in correlacoes.values()
        if info.get("grupo_id") == grupo_id
    })
    print(f"  [BANCO_COORDS] correlação manual: {len(pares)} endereço(s) → grupo {grupo_id!r} ({user_id or '?'})")
    return {"ok": True, "grupo_id": grupo_id, "membros": membros}


def banco_coords_aplicar(rows: list, user_id: str = "") -> list:
    banco = banco_coords_carregar()
    overrides_usuario = banco["overrides"].get(user_id, {}) if user_id else {}
    globais = banco["global"]
    correlacoes = banco.get("correlacoes", {})
    if not overrides_usuario and not globais and not correlacoes:
        return rows
    for row in rows:
        chave = _normalizar_endereco(row.get("address", ""))
        entrada = overrides_usuario.get(chave)
        fonte = "override"
        if not entrada:
            entrada = globais.get(chave)
            fonte = entrada.get("fonte", "banco") if entrada else None
        if entrada:
            lat = str(entrada["lat"])
            lon = str(entrada["lon"])
            row["lat"]         = lat
            row["lon"]         = lon
            row["coord"]       = lat + "," + lon
            row["do_banco"]    = True
            row["fonte_coord"] = fonte
        elif correlacoes:
            # Sem coordenada própria (nem override, nem global) — mas talvez
            # um "parente" correlacionado manualmente já tenha uma. Isso é só
            # uma dica (correlacao_sugerida) pro front avisar o usuário; não
            # aplica a coordenada sozinho.
            info = correlacoes.get(chave)
            if info:
                grupo_id = info["grupo_id"]
                for c2, i2 in correlacoes.items():
                    if c2 == chave or i2.get("grupo_id") != grupo_id:
                        continue
                    sugestao = overrides_usuario.get(c2) or globais.get(c2)
                    if sugestao:
                        row["correlacao_sugerida"] = {
                            "endereco": i2["endereco_original"],
                            "lat": sugestao["lat"],
                            "lon": sugestao["lon"],
                        }
                        break
    return rows

# ─── Geocodificação sob demanda (confirmação na tela de loading) ──────

_RE_QUADRA_LOTE     = re.compile(r"\b(qd|quadra|lt|lote|q\d+|l\d+)\b", re.IGNORECASE)
_RE_NUMERO_COMPOSTO = re.compile(r"\d+\s*[-/]\s*\d+")
_RE_RUA_CODIGO      = re.compile(r"\b[A-Za-z]\s?\d{2,}\b")

def endereco_e_simples(endereco: str) -> bool:
    """
    Regra combinada com o Geni: endereço de rua com número simples
    (ex.: "RUA D, 385") x endereço com quadra/lote ou número composto
    (ex.: "RUA C152, 343-1"). Mantido pra outras decisões do fluxo; não
    decide mais nada sobre chamar o Google — isso saiu do fluxo
    automático (ver geocode_confirmar).
    """
    if not endereco:
        return False
    if _RE_QUADRA_LOTE.search(endereco):
        return False
    if _RE_NUMERO_COMPOSTO.search(endereco):
        return False
    if _RE_RUA_CODIGO.search(endereco):
        return False
    return True

# O endereço já chega padronizado do tratamento_dados.py como
# "RUA X, qd-lt", "RUA X, num" ou "RUA X, ED. Nome" (ou só "RUA X", sem
# nada). A validação estrita por resultType/houseNumber que tinha aqui
# foi removida — estava rejeitando muito endereço válido (o HERE nem
# sempre marca resultType como houseNumber mesmo quando acerta). Agora
# aceita o primeiro resultado do HERE direto, como antes.

def _here_geocode_query(query: str) -> tuple[float, float] | None:
    if not HERE_API_KEY:
        return None
    try:
        r = requests.get(
            "https://geocode.search.hereapi.com/v1/geocode",
            params={
                "q": f"{query}, {HERE_CIDADE_UF}",
                "apiKey": HERE_API_KEY,
                "limit": 1,
                "lang": "pt-BR",
            },
            timeout=8,
        )
        r.raise_for_status()
        itens = r.json().get("items") or []
        if not itens:
            return None
        pos = itens[0].get("position") or {}
        lat, lon = pos.get("lat"), pos.get("lng")
        if lat is None or lon is None:
            return None
        return round(float(lat), 6), round(float(lon), 6)
    except Exception as e:
        print(f"  [GEOCODE/HERE] falha em {query!r}: {e}")
        return None

def _google_geocode_query(query: str) -> tuple[float, float] | None:
    if not GOOGLE_GEOCODING_API_KEY:
        return None
    try:
        r = requests.get(
            GOOGLE_GEOCODING_URL,
            params={
                "address": f"{query}, {HERE_CIDADE_UF}",
                "key": GOOGLE_GEOCODING_API_KEY,
                "language": "pt-BR",
                "region": "br",
            },
            timeout=8,
        )
        r.raise_for_status()
        body = r.json()
        if body.get("status") != "OK":
            return None
        resultados = body.get("results") or []
        if not resultados:
            return None
        loc = (resultados[0].get("geometry") or {}).get("location") or {}
        lat, lon = loc.get("lat"), loc.get("lng")
        if lat is None or lon is None:
            return None
        return round(float(lat), 6), round(float(lon), 6)
    except Exception as e:
        print(f"  [GEOCODE/GOOGLE] falha em {query!r}: {e}")
        return None

# ═══════════════════════════════════════════════════════════════════
#  USUÁRIOS
# ═══════════════════════════════════════════════════════════════════

def _hash_senha(senha: str) -> str:
    return hashlib.sha256(senha.encode("utf-8")).hexdigest()

def carregar_usuarios() -> dict:
    return _carregar_com_recuperacao(USERS_FILE, {})

def salvar_usuarios(users: dict):
    _salvar_com_backup(
        USERS_FILE,
        json.dumps(users, ensure_ascii=False, indent=2),
        dados_novos_vazios=(len(users) == 0),
    )

def _buscar_usuario(users: dict, username: str):
    alvo = username.strip().lower()
    for chave, dados in users.items():
        if chave.lower() == alvo:
            return chave, dados
    return None, None

def autenticar_usuario(username: str, senha: str) -> tuple[str, str] | None:
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None or u.get("hash") != _hash_senha(senha):
        return None
    if not u.get("id"):
        u["id"] = str(uuid.uuid4())
        salvar_usuarios(users)
    return u["id"], chave

def usuario_e_admin(username: str) -> bool:
    users = carregar_usuarios()
    _, u = _buscar_usuario(users, username)
    return bool(u and u.get("is_admin"))

def usuario_tem_acesso_ativo(username: str) -> bool:
    users = carregar_usuarios()
    _, u = _buscar_usuario(users, username)
    if u is None:
        return False
    expira_raw = u.get("acesso_expira_em")
    if expira_raw:
        try:
            if datetime.now() < datetime.fromisoformat(expira_raw):
                return True
        except ValueError:
            pass
    return int(u.get("avulsa_creditos", 0) or 0) > 0

def _contagem_hoje(u: dict) -> int:
    hoje = datetime.now().strftime("%Y-%m-%d")
    c = u.get("importacoes_hoje", {})
    if not isinstance(c, dict) or c.get("data") != hoje:
        return 0
    return int(c.get("count", 0) or 0)

def usuario_pode_importar_hoje(username: str) -> tuple[bool, str]:
    users = carregar_usuarios()
    _, u = _buscar_usuario(users, username)
    if u is None:
        return False, "Usuário não encontrado."
    plano_ativo = u.get("plano_ativo")
    if not plano_ativo:
        return True, ""
    limite = PLANOS.get(plano_ativo, {}).get("importacoes_por_dia")
    if limite is None:
        return True, ""
    usadas = _contagem_hoje(u)
    if usadas >= limite:
        sufixo = "ão" if limite == 1 else "ões"
        return False, (f"Limite diário do seu plano atingido "
                       f"({usadas}/{limite} importaç{sufixo} hoje). Volte amanhã.")
    return True, ""

def registrar_importacao_hoje(username: str):
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None or not u.get("plano_ativo"):
        return
    hoje = datetime.now().strftime("%Y-%m-%d")
    c = u.get("importacoes_hoje", {})
    if not isinstance(c, dict) or c.get("data") != hoje:
        c = {"data": hoje, "count": 0}
    c["count"] = int(c.get("count", 0) or 0) + 1
    u["importacoes_hoje"] = c
    salvar_usuarios(users)
    limite = PLANOS.get(u["plano_ativo"], {}).get("importacoes_por_dia")
    print(f"  [LIMITE] {chave}: {c['count']} importação(ões) hoje (limite: {limite}).")

def usuario_consumir_credito_avulso_se_necessario(username: str):
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None:
        return
    tem_acesso_mensal = False
    expira_raw = u.get("acesso_expira_em")
    if expira_raw:
        try:
            tem_acesso_mensal = datetime.now() < datetime.fromisoformat(expira_raw)
        except ValueError:
            pass
    if tem_acesso_mensal:
        return
    creditos = int(u.get("avulsa_creditos", 0) or 0)
    if creditos > 0:
        u["avulsa_creditos"] = creditos - 1
        salvar_usuarios(users)
        print(f"  [ASSINATURA] Crédito avulso consumido por \"{chave}\" (restam {creditos - 1}).")

# ═══════════════════════════════════════════════════════════════════
#  ADMIN
# ═══════════════════════════════════════════════════════════════════

def admin_listar_usuarios() -> list:
    users = carregar_usuarios()
    # Stat de precisão de coordenadas — só o admin vê (ver
    # banco_coords_promover_manual pra como confirmadas/corrigidas são contadas).
    stats_coords = banco_coords_carregar().get("stats_usuario", {})
    out = []
    for nome, dados in users.items():
        s = stats_coords.get(dados.get("id"), {})
        out.append({
            "usuario":             nome,
            "email":               dados.get("email", ""),
            "telefone":            dados.get("telefone", ""),
            "is_admin":            bool(dados.get("is_admin", False)),
            "acesso_expira_em":    dados.get("acesso_expira_em"),
            "origem_acesso":       dados.get("origem_acesso"),
            "avulsa_creditos":     int(dados.get("avulsa_creditos", 0) or 0),
            "plano_solicitado":    dados.get("plano_solicitado"),
            "plano_solicitado_em": dados.get("plano_solicitado_em"),
            "coords_confirmadas":  s.get("confirmadas", 0),
            "coords_corrigidas":   s.get("corrigidas", 0),
        })
    out.sort(key=lambda u: u["usuario"].lower())
    return out

def admin_criar_usuario(username: str, senha: str, email: str = "", is_admin: bool = False) -> tuple[bool, str]:
    username = (username or "").strip()
    email    = (email or "").strip().lower()
    if not username or len(username) < 3:
        return False, "Usuário deve ter pelo menos 3 caracteres."
    if not senha or len(senha) < 4:
        return False, "Senha deve ter pelo menos 4 caracteres."
    if email and not _email_valido(email):
        return False, "Email inválido."
    users = carregar_usuarios()
    chave_existente, _ = _buscar_usuario(users, username)
    if chave_existente is not None:
        return False, "Usuário já existe."
    if email and any(u.get("email", "").lower() == email for u in users.values()):
        return False, "Este email já está cadastrado em outra conta."
    novo = {"id": str(uuid.uuid4()), "hash": _hash_senha(senha)}
    if email:
        novo["email"] = email
    if is_admin:
        novo["is_admin"] = True
    else:
        # Mesma regra do cadastro público: toda conta nova (que não seja
        # admin) começa com 3 dias grátis automáticos.
        _conceder_teste_gratis(novo)
    users[username] = novo
    salvar_usuarios(users)
    return True, "Usuário criado com sucesso."

def admin_resetar_senha(username: str, nova_senha: str) -> tuple[bool, str]:
    if not nova_senha or len(nova_senha) < 4:
        return False, "A nova senha deve ter pelo menos 4 caracteres."
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None:
        return False, "Usuário não encontrado."
    u["hash"] = _hash_senha(nova_senha)
    salvar_usuarios(users)
    return True, "Senha redefinida com sucesso."

def admin_editar_contato(username: str, email: str = "", telefone: str = "") -> tuple[bool, str]:
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None:
        return False, "Usuário não encontrado."
    email = (email or "").strip().lower()
    if email and not _email_valido(email):
        return False, "Email inválido."
    if email and any(k.lower() != chave.lower() and d.get("email", "").lower() == email
                     for k, d in users.items()):
        return False, "Este email já está cadastrado em outra conta."
    telefone_norm = _normalizar_telefone(telefone or "")
    if telefone_norm and not _telefone_valido(telefone_norm):
        return False, "Telefone inválido. Use DDD + 9 + número (ex: 62 9 91153473)."
    if email:
        u["email"] = email
    else:
        u.pop("email", None)
    if telefone_norm:
        u["telefone"] = telefone_norm
    else:
        u.pop("telefone", None)
    salvar_usuarios(users)
    return True, "Dados atualizados com sucesso."

def admin_apagar_usuario(username: str, quem_pediu: str) -> tuple[bool, str]:
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None:
        return False, "Usuário não encontrado."
    if chave.lower() == (quem_pediu or "").strip().lower():
        return False, "Você não pode apagar sua própria conta de admin enquanto está logado nela."
    del users[chave]
    salvar_usuarios(users)
    user_id = u.get("id")
    if user_id:
        for t in [t for t, s in _sessoes.items() if s.get("user_id") == user_id]:
            del _sessoes[t]
    return True, "Usuário apagado com sucesso."

def admin_liberar_acesso(username: str, dias: int) -> tuple[bool, str]:
    try:
        dias = int(dias)
    except (TypeError, ValueError):
        return False, "Número de dias inválido."
    if dias < 1:
        return False, "Informe pelo menos 1 dia de acesso."
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None:
        return False, "Usuário não encontrado."
    expira_em = datetime.now() + timedelta(days=dias)
    u["acesso_expira_em"] = expira_em.isoformat()
    u.pop("plano_ativo", None)
    u.pop("origem_acesso", None)
    salvar_usuarios(users)
    return True, f"Acesso liberado até {expira_em.strftime('%d/%m/%Y %H:%M')}."

def admin_limpar_cache_usuario(username: str) -> tuple[bool, str]:
    """Limpeza pontual manual (disparada pelo admin, não automática a cada
    deploy): apaga o histórico de rotas salvas desse usuário — força
    reimportar do zero da próxima vez, resolvendo o caso de uma rota
    antiga ficar com coordenada desatualizada/errada — e marca
    cache_limpo_em, que o frontend usa (via /auth/status) pra também
    descartar o progresso do Modo de Entrega salvo no navegador dele.
    Nunca mexe em gamificação/XP nem no banco de coordenadas confirmadas
    (só nos overrides pessoais dele, mesmo comportamento de quando uma
    rota sai do histórico normalmente)."""
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None:
        return False, "Usuário não encontrado."
    user_id = u.get("id", "")

    historico = carregar_historico()
    antes = len(historico)
    novo_historico = [h for h in historico if h.get("user_id") != user_id]
    # forcar=True: é intencional o histórico desse usuário zerar por completo
    # (ex.: só ele tinha rota salva) — a confirmação já foi pedida no
    # frontend antes de chamar essa função, não é acidente pra bloquear.
    salvar_historico(novo_historico, forcar=True)
    removidos = antes - len(novo_historico)

    banco_coords_limpar_overrides_usuario(user_id)

    u["cache_limpo_em"] = datetime.now().isoformat()
    salvar_usuarios(users)

    return True, (
        f'Histórico limpo ({removidos} rota(s) removida(s)) para "{chave}". '
        'O progresso do Modo de Entrega salvo no navegador dele também será limpo no próximo acesso.'
    )

def admin_limpar_cache_todos_usuarios() -> tuple[bool, str]:
    """Igual admin_limpar_cache_usuario, mas pra todos os usuários de uma
    vez — zera o histórico de rotas inteiro (todo mundo precisa reimportar
    na próxima vez) e os overrides pessoais de todo mundo no banco de
    coordenadas, e marca cache_limpo_em em cada usuário. Nunca mexe em
    gamificação/XP nem no "global" do banco de coordenadas confirmadas
    (só o "overrides", que é sempre pessoal/temporário por natureza)."""
    historico = carregar_historico()
    total_rotas = len(historico)
    salvar_historico([], forcar=True)

    banco = banco_coords_carregar()
    banco["overrides"] = {}
    banco_coords_salvar(banco)

    users = carregar_usuarios()
    agora = datetime.now().isoformat()
    for u in users.values():
        u["cache_limpo_em"] = agora
    salvar_usuarios(users)

    return True, (
        f"Histórico limpo para todos os usuários ({total_rotas} rota(s) removida(s) no total). "
        "O progresso do Modo de Entrega salvo no navegador de cada um será limpo no próximo acesso."
    )

def admin_revogar_acesso(username: str) -> tuple[bool, str]:
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None:
        return False, "Usuário não encontrado."
    u.pop("acesso_expira_em", None)
    u.pop("plano_ativo", None)
    u.pop("origem_acesso", None)
    salvar_usuarios(users)
    return True, "Acesso revogado."

# ─── Planos ────────────────────────────────────────────────────────

def _creditar_plano(u: dict, plano_id: str) -> str:
    plano = PLANOS[plano_id]
    if plano["tipo"] == "avulso":
        u["avulsa_creditos"] = int(u.get("avulsa_creditos", 0) or 0) + 1
        u.pop("plano_ativo", None)
        u.pop("origem_acesso", None)
        return "1 crédito de importação avulsa liberado."
    expira_em = datetime.now() + timedelta(days=plano["dias"])
    u["acesso_expira_em"] = expira_em.isoformat()
    u["plano_ativo"] = plano_id
    u.pop("origem_acesso", None)
    return f"{plano['nome']} liberado até {expira_em.strftime('%d/%m/%Y %H:%M')}."

def usuario_solicitar_plano(username: str, plano_id: str) -> tuple[bool, str]:
    plano = PLANOS.get(plano_id)
    if plano is None:
        return False, "Plano inválido."
    if plano.get("pagamento_automatico"):
        return False, f"{plano['nome']} usa pagamento automático — use o botão de pagamento."
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None:
        return False, "Usuário não encontrado."
    u["plano_solicitado"]    = plano_id
    u["plano_solicitado_em"] = datetime.now().isoformat()
    u.pop("pagamento_pendente", None)
    salvar_usuarios(users)
    return True, f"Solicitação de {plano['nome']} registrada. Aguarde a liberação do administrador."

def admin_confirmar_plano(username: str) -> tuple[bool, str]:
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None:
        return False, "Usuário não encontrado."
    plano_id = u.get("plano_solicitado")
    plano    = PLANOS.get(plano_id)
    if plano is None:
        return False, "Este usuário não tem solicitação de plano pendente."
    _creditar_plano(u, plano_id)
    msg = (f"1 crédito de importação avulsa liberado para \"{chave}\"."
           if plano["tipo"] == "avulso"
           else f"{plano['nome']} liberado para \"{chave}\".")
    u.pop("plano_solicitado", None)
    u.pop("plano_solicitado_em", None)
    salvar_usuarios(users)
    return True, msg

def admin_rejeitar_plano(username: str) -> tuple[bool, str]:
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None:
        return False, "Usuário não encontrado."
    u.pop("plano_solicitado", None)
    u.pop("plano_solicitado_em", None)
    salvar_usuarios(users)
    return True, "Solicitação removida."

# ═══════════════════════════════════════════════════════════════════
#  INFINITEPAY
# ═══════════════════════════════════════════════════════════════════

def _infinitepay_gerar_link(order_nsu: str, plano: dict, redirect_url: str, webhook_url: str) -> tuple[bool, str]:
    payload = {
        "handle":       INFINITEPAY_HANDLE,
        "redirect_url": redirect_url,
        "webhook_url":  webhook_url,
        "order_nsu":    order_nsu,
        "items": [{"quantity": 1, "price": int(round(plano["preco"] * 100)), "description": plano["nome"]}],
    }
    try:
        resp = requests.post(INFINITEPAY_LINKS_URL, json=payload, timeout=10)
        data = resp.json()
    except Exception as e:
        return False, f"Não foi possível conectar à InfinitePay: {e}"
    url = data.get("url")
    if not resp.ok or not url:
        erro = data.get("message") or data.get("error") or f"Erro {resp.status_code} ao gerar o link de pagamento."
        return False, erro
    return True, url

def infinitepay_consultar_pagamento(order_nsu: str, transaction_nsu: str = "", slug: str = "") -> tuple[bool, dict | str]:
    payload = {"handle": INFINITEPAY_HANDLE, "order_nsu": order_nsu,
               "transaction_nsu": transaction_nsu, "slug": slug}
    try:
        resp = requests.post(INFINITEPAY_PAYMENT_CHECK_URL, json=payload, timeout=10)
        data = resp.json()
    except Exception as e:
        return False, f"Não foi possível consultar o pagamento: {e}"
    if not resp.ok:
        return False, data.get("message") or f"Erro {resp.status_code} ao consultar pagamento."
    return True, data

def usuario_iniciar_pagamento(username: str, plano_id: str, base_url: str) -> tuple[bool, str]:
    plano = PLANOS.get(plano_id)
    if plano is None:
        return False, "Plano inválido."
    if not plano.get("pagamento_automatico"):
        return False, f"{plano['nome']} ainda usa o fluxo de solicitação manual."
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, username)
    if u is None:
        return False, "Usuário não encontrado."
    pendente = u.get("pagamento_pendente")
    if pendente and pendente.get("plano_id") == plano_id:
        try:
            criado_em = datetime.fromisoformat(pendente["criado_em"])
            if datetime.now() - criado_em < timedelta(minutes=PAGAMENTO_LINK_REUSE_MINUTOS):
                return True, pendente["url"]
        except (KeyError, ValueError):
            pass
    order_nsu    = uuid.uuid4().hex
    redirect_url = f"{base_url}/?pagamento=retorno"
    webhook_url  = f"{base_url}/webhook/infinitepay"
    ok, resultado = _infinitepay_gerar_link(order_nsu, plano, redirect_url, webhook_url)
    if not ok:
        return False, resultado
    u["pagamento_pendente"] = {
        "plano_id":  plano_id,
        "order_nsu": order_nsu,
        "url":       resultado,
        "criado_em": datetime.now().isoformat(),
    }
    salvar_usuarios(users)
    return True, resultado

def processar_pagamento_confirmado(order_nsu: str, transaction_nsu: str = "", receipt_url: str = "") -> tuple[bool, str]:
    users = carregar_usuarios()
    chave_alvo, u_alvo = None, None
    for chave, u in users.items():
        pendente = u.get("pagamento_pendente")
        if pendente and pendente.get("order_nsu") == order_nsu:
            chave_alvo, u_alvo = chave, u
            break
    if u_alvo is None:
        return False, "Pedido não encontrado."
    plano_id = u_alvo["pagamento_pendente"].get("plano_id")
    if plano_id not in PLANOS:
        return False, "Plano do pedido não existe mais."
    detalhe = _creditar_plano(u_alvo, plano_id)
    u_alvo.pop("pagamento_pendente", None)
    u_alvo["ultimo_pagamento"] = {
        "plano_id": plano_id, "order_nsu": order_nsu,
        "transaction_nsu": transaction_nsu, "receipt_url": receipt_url,
        "pago_em": datetime.now().isoformat(),
    }
    salvar_usuarios(users)
    print(f"  [INFINITEPAY] Pagamento confirmado para \"{chave_alvo}\" (order_nsu={order_nsu}). {detalhe}")
    return True, detalhe

# ═══════════════════════════════════════════════════════════════════
#  LER PLANILHA PROCESSADA
# ═══════════════════════════════════════════════════════════════════

def ler_processado(user_id: str = "", caminho: "Path | None" = None):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl não instalado.")
    if caminho is not None:
        path = Path(caminho)
    else:
        path = Path(ARQ_VALIDADO) if Path(ARQ_VALIDADO).exists() else Path(ARQ_PROCESSADO)
    if not path.exists():
        raise FileNotFoundError(f"{path} não encontrado.")
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]

    def find_col(pats):
        for pat in pats:
            for i, h in enumerate(headers):
                if re.search(pat, h, re.IGNORECASE):
                    return i
        return None

    col_addr    = find_col([r"destination.?address", r"reformado"])
    col_stop    = find_col([r"sequence", r"stop", r"seq"])
    col_lat     = find_col([r"\blatitude\b", r"\blat\b"])
    col_lon     = find_col([r"\blongitude\b", r"\blon\b", r"\blng\b"])
    col_coord   = find_col([r"coordenadas", r"coord"])
    col_count   = find_col([r"rotas_iguais"])
    col_stops   = find_col([r"stops do grupo"])
    col_orig    = find_col([r"endere.o_original", r"original"])
    col_membros = find_col([r"membros.?json", r"membros"])
    col_validacao_here = find_col([r"validacao_here", r"validacao.here"])
    col_contato = find_col([r"\bcontato\b", r"contact", r"telefone", r"\bphone\b"])
    col_bairro  = find_col([r"\bbairro\b", r"neighbo"])
    col_zip     = find_col([r"zip", r"postal", r"\bcep\b"])

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        def g(idx):
            if idx is None or idx >= len(row):
                return ""
            v = row[idx]
            return str(v).strip() if v is not None else ""

        lat   = g(col_lat)
        lon   = g(col_lon)
        coord = g(col_coord) or (f"{lat},{lon}" if lat and lon else "")
        count = int(g(col_count) or 1)
        endereco_original = g(col_orig)

        contato_grupo = g(col_contato)

        membros = []
        membros_raw = g(col_membros)
        if membros_raw:
            try:
                parsed = json.loads(membros_raw)
                if isinstance(parsed, list):
                    membros = [
                        {
                            "stop": str(m.get("stop", "")),
                            "original": str(m.get("original", "")),
                            "contato": str(m.get("contato", "")),
                        }
                        for m in parsed if isinstance(m, dict)
                    ]
            except Exception:
                membros = []

        if not membros:
            seq_fallback   = [s.strip() for s in g(col_stop).split(",") if s.strip()]
            stops_fallback = seq_fallback or [s.strip() for s in g(col_stops).replace("Stop:", "").split(",") if s.strip()]
            origs_fallback = [o.strip() for o in endereco_original.split("|") if o.strip()]
            n = max(len(stops_fallback), len(origs_fallback), 1)
            membros = [
                {"stop": stops_fallback[i] if i < len(stops_fallback) else "",
                 "original": origs_fallback[i] if i < len(origs_fallback) else endereco_original,
                 "contato": contato_grupo}
                for i in range(n)
            ]

        entry = {
            "address":            g(col_addr),
            "stop":               g(col_stop),
            "lat":                lat,
            "lon":                lon,
            "coord":              coord,
            "group_size":         count,
            "group_stops":        g(col_stops),
            "group_label":        g(col_addr),
            "endereco_original":  endereco_original,
            "membros":            membros,
            "contato":            contato_grupo,
            "bairro":             g(col_bairro),
            "zip":                g(col_zip),
            "validacao_here":     g(col_validacao_here),
            "_cid":               str(uuid.uuid4()),
        }
        rows.append(entry)

    rows = banco_coords_aplicar(rows, user_id)
    return rows, headers

# ═══════════════════════════════════════════════════════════════════
#  OTIMIZAÇÃO DE ROTA (OSRM /trip/)
# ═══════════════════════════════════════════════════════════════════

def _osrm_otimizar_sequencia(coords: list[tuple[float, float]]) -> list[int] | None:
    """
    Recebe uma lista de (lat, lon) e devolve a ordem otimizada dos ÍNDICES
    originais (mesmo tamanho da entrada), usando o serviço público OSRM
    /trip/ (baseado em OpenStreetMap). O primeiro ponto da lista é fixado
    como início da rota (source=first) e não faz retorno ao ponto de
    partida (roundtrip=false).

    Retorna None se o serviço falhar ou a resposta vier inconsistente —
    nesse caso o chamador deve manter a ordem original.
    """
    if len(coords) < 2:
        return list(range(len(coords)))

    coord_str = ";".join(f"{lon:.6f},{lat:.6f}" for lat, lon in coords)
    url = f"{OSRM_BASE_URL}/trip/v1/driving/{coord_str}"
    params = {"source": "first", "roundtrip": "false", "overview": "false"}
    # O servidor de demonstração do OSRM fica atrás de uma proteção anti-bot
    # que costuma barrar requisições com o User-Agent padrão do `requests`
    # (ex: "python-requests/2.x"). Um UA de navegador evita esse bloqueio.
    headers = {
        "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/124.0.0.0 Safari/537.36"),
        "Accept": "application/json",
    }

    try:
        r = requests.get(url, params=params, headers=headers, timeout=30)
        status = r.status_code
        if status != 200:
            print(f"  [OSRM] HTTP {status} — corpo: {r.text[:300]!r}")
            return None
        data = r.json()
    except requests.exceptions.RequestException as e:
        print(f"  [OSRM] falha na chamada: {type(e).__name__}: {e}")
        return None

    if data.get("code") != "Ok":
        print(f"  [OSRM] resposta inesperada: {data.get('code')} — {data.get('message', '')}")
        return None

    waypoints = data.get("waypoints", [])
    if len(waypoints) != len(coords):
        print("  [OSRM] número de waypoints não bate com a entrada.")
        return None

    # waypoints[i]['waypoint_index'] = posição do ponto i na rota otimizada
    try:
        ordem = sorted(range(len(coords)), key=lambda i: waypoints[i]["waypoint_index"])
    except (KeyError, TypeError):
        return None
    return ordem

def _ors_otimizar_sequencia(coords: list[tuple[float, float]]) -> list[int] | None:
    """
    Mesmo contrato de _osrm_otimizar_sequencia: recebe (lat, lon) e devolve a
    ordem otimizada dos ÍNDICES originais. Usa o endpoint /optimization do
    OpenRouteService (motor VROOM), que é hospedado com SLA — muito mais
    confiável que o servidor público de demonstração do OSRM.

    Fixa o primeiro ponto da lista como início do "veículo" (não é tratado
    como job) e todos os demais pontos entram como jobs a visitar na ordem
    mais eficiente. Não exige retorno ao ponto de partida (sem "end").
    """
    if len(coords) < 2:
        return list(range(len(coords)))
    if not ORS_API_KEY:
        print("  [ORS] ORS_API_KEY não configurada — pulei para o próximo serviço.")
        return None

    inicio = coords[0]
    jobs = [
        {"id": i, "location": [lon, lat]}
        for i, (lat, lon) in enumerate(coords) if i != 0
    ]
    payload = {
        "jobs": jobs,
        "vehicles": [{
            "id": 1,
            "profile": "driving-car",
            "start": [inicio[1], inicio[0]],
        }],
    }
    headers = {
        "Authorization": ORS_API_KEY,
        "Content-Type": "application/json",
        "Accept": "application/json",
    }

    try:
        r = requests.post(f"{ORS_BASE_URL}/optimization", json=payload, headers=headers, timeout=30)
        if r.status_code != 200:
            print(f"  [ORS] HTTP {r.status_code} — corpo: {r.text[:300]!r}")
            return None
        data = r.json()
    except requests.exceptions.RequestException as e:
        print(f"  [ORS] falha na chamada: {type(e).__name__}: {e}")
        return None

    routes = data.get("routes") or []
    if not routes:
        print(f"  [ORS] resposta sem rotas: {str(data)[:300]}")
        return None

    steps = routes[0].get("steps", [])
    ordem_jobs = [s["id"] for s in steps if s.get("type") == "job"]
    if len(ordem_jobs) != len(jobs):
        print(f"  [ORS] {len(ordem_jobs)} jobs retornados, esperava {len(jobs)}.")
        return None

    return [0] + ordem_jobs

def _otimizar_sequencia(coords: list[tuple[float, float]]) -> tuple[list[int] | None, str]:
    """
    Escolhe o serviço de otimização: OpenRouteService primeiro (se a chave
    estiver configurada — é o caminho recomendado e mais estável), caindo
    para o OSRM público como último recurso caso a chave não esteja
    configurada ou o ORS falhe. Retorna (ordem, nome_do_servico_usado).
    """
    if ORS_API_KEY:
        ordem = _ors_otimizar_sequencia(coords)
        if ordem is not None:
            return ordem, "openrouteservice"
        print("  [OTIMIZAR] ORS falhou, tentando OSRM público como fallback...")

    ordem = _osrm_otimizar_sequencia(coords)
    return ordem, "osrm"

# ═══════════════════════════════════════════════════════════════════
#  BOOTSTRAP DO ADMIN
# ═══════════════════════════════════════════════════════════════════

def _bootstrap_admin():
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, "admin")
    if u is None:
        senha_inicial = os.environ.get("ADMIN_PASS", "admin123")
        users["admin"] = {
            "id":       str(uuid.uuid4()),
            "hash":     _hash_senha(senha_inicial),
            "is_admin": True,
        }
        salvar_usuarios(users)
        print(f"  [ADMIN] Usuário 'admin' criado. Senha inicial: {senha_inicial!r}")
    elif not u.get("is_admin"):
        u["is_admin"] = True
        salvar_usuarios(users)
        print("  [ADMIN] Usuário 'admin' recebeu a flag is_admin.")

# ═══════════════════════════════════════════════════════════════════
#  ROTAS  ──  GET
# ═══════════════════════════════════════════════════════════════════

@app.get("/ping")
async def ping():
    return {"ok": True}

@app.get("/")
@app.get("/index")
@app.get(f"/{HTML_FILE}")
async def serve_html():
    html_path = Path(HTML_FILE)
    if not html_path.exists():
        raise HTTPException(status_code=404, detail="rota_manager1.html não encontrado.")
    return FileResponse(html_path, media_type="text/html; charset=utf-8")

@app.get("/dados")
async def get_dados(request: Request):
    sess = _sessao_ou_401(request)
    if sess["dados"] is None:
        return err_json("Nenhum dado processado ainda.", 404)
    rows, headers = sess["dados"]
    return ok_json({"ok": True, "arquivo": ARQ_PROCESSADO, "rows": rows, "headers": headers})

@app.get("/auth/status")
async def auth_status(request: Request):
    sess = _sessao_ou_401(request)
    is_admin = bool(sess.get("is_admin"))
    tem_acesso = is_admin or usuario_tem_acesso_ativo(sess["usuario"])

    users = carregar_usuarios()
    _, u = _buscar_usuario(users, sess["usuario"])
    # Marcado por admin_limpar_cache_usuario — o frontend compara com o
    # último valor que já viu e, se for mais novo, descarta o progresso do
    # Modo de Entrega salvo no localStorage dele (ver
    # _aplicarLimpezaCacheSeNecessario no rota_manager1.html).
    cache_limpo_em = (u or {}).get("cache_limpo_em")

    # Aviso de teste grátis acabando — só pra quem ainda está no
    # "teste_gratis" (não afeta plano pago nem créditos avulsos).
    aviso_trial = None
    if not is_admin and tem_acesso and u and u.get("origem_acesso") == "teste_gratis":
        expira_raw = u.get("acesso_expira_em")
        if expira_raw:
            try:
                expira_em = datetime.fromisoformat(expira_raw)
                horas_restantes = (expira_em - datetime.now()).total_seconds() / 3600
                if horas_restantes <= AVISO_TRIAL_HORAS_ANTES:
                    aviso_trial = {
                        "dias_restantes":  max(0, round(horas_restantes / 24, 1)),
                        "expira_em":       expira_raw,
                    }
            except ValueError:
                pass

    return ok_json({
        "ok": True, "tem_acesso": tem_acesso, "is_admin": is_admin,
        "aviso_trial": aviso_trial, "cache_limpo_em": cache_limpo_em,
    })

@app.get("/api/config")
async def api_config(request: Request):
    """Chaves de mapa/geocoding pro frontend (popup de correção, mapa de
    rotas, scan/OCR) — liberadas só pra quem está logado, nunca hardcoded
    no HTML/JS servido."""
    _sessao_ou_401(request)
    return ok_json({
        "ok": True,
        "hereApiKey": HERE_API_KEY,
        "googleMapsKey": GOOGLE_MAPS_BROWSER_KEY,
        "mapboxToken": MAPBOX_BROWSER_TOKEN,
    })

@app.get("/api/perfil/me")
async def perfil_gamificacao_me(request: Request):
    """Atalho pro perfil de gamificação do usuário logado (sem precisar
    que o frontend conheça o user_id — só o token de sessão)."""
    sess = _sessao_ou_401(request)
    perfil = _garantir_perfil(sess["user_id"])
    return ok_json({
        "ok": True,
        "perfil": perfil.model_dump(),
        "nome_exibicao": perfil.nome_personagem or sess["usuario"],
        "xp_para_proximo_nivel": gamification.xp_necessario_para_nivel(perfil.nivel),
        "icon_base_path": gamification.ICON_BASE_PATH,
        "badges_desbloqueadas": gamification.calcular_badges_desbloqueadas(perfil.nivel),
        "itens_desbloqueados": gamification.calcular_itens_desbloqueados(perfil),
    })

@app.post("/api/perfil/me/nome")
async def perfil_gamificacao_atualizar_nome(request: Request):
    """Define/limpa o apelido do personagem do usuário logado. Se vazio,
    o personagem volta a usar o nome de usuário normal."""
    sess = _sessao_ou_401(request)
    data = await request.json()
    nome_novo = (data.get("nome_personagem") or "").strip()
    if nome_novo:
        if len(nome_novo) < 2 or len(nome_novo) > 20:
            return err_json("O nome do personagem deve ter entre 2 e 20 caracteres.")
    perfil = _garantir_perfil(sess["user_id"])
    perfil.nome_personagem = nome_novo or None
    gamification.salvar_usuario(perfil)
    return ok_json({
        "ok": True,
        "msg": "Nome do personagem atualizado.",
        "nome_exibicao": perfil.nome_personagem or sess["usuario"],
    })

@app.get("/api/perfil/ranking")
async def perfil_ranking(request: Request):
    """Ranking de todos os jogadores por nível (desempate por XP total)."""
    sess = _sessao_ou_401(request)
    users = carregar_usuarios()
    id_para_usuario = {u.get("id"): chave for chave, u in users.items() if u.get("id")}

    ranking = []
    for p in gamification.listar_todos_perfis():
        username = id_para_usuario.get(p.user_id, "Piloto")
        ranking.append({
            "user_id":       p.user_id,
            "nome_exibicao": p.nome_personagem or username,
            "nivel":         p.nivel,
            "xp_total":      p.xp_total,
        })
    ranking.sort(key=lambda r: (-r["nivel"], -r["xp_total"]))
    for i, r in enumerate(ranking, start=1):
        r["posicao"] = i

    minha_posicao = next((r for r in ranking if r["user_id"] == sess["user_id"]), None)
    return ok_json({
        "ok": True,
        "ranking": ranking[:20],
        "minha_posicao": minha_posicao,
        "total_jogadores": len(ranking),
    })

@app.get("/auth/perfil")
async def auth_perfil(request: Request):
    sess = _sessao_ou_401(request)
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, sess["usuario"])
    if u is None:
        return err_json("Usuário não encontrado.")
    return ok_json({"ok": True, "usuario": chave, "email": u.get("email", ""), "telefone": u.get("telefone", "")})

@app.get("/planos")
async def get_planos(request: Request):
    _sessao_ou_401(request)
    planos = [{"id": pid, **dados} for pid, dados in PLANOS.items()]
    return ok_json({"ok": True, "planos": planos})

@app.get("/assinatura/status")
async def assinatura_status(request: Request):
    sess = _sessao_ou_401(request)
    users = carregar_usuarios()
    _, u = _buscar_usuario(users, sess["usuario"])
    if u is None:
        return err_json("Usuário não encontrado.", 404)
    plano_solicitado = u.get("plano_solicitado")
    pendente = u.get("pagamento_pendente")
    return ok_json({
        "ok":                    True,
        "acesso_expira_em":      u.get("acesso_expira_em"),
        "origem_acesso":         u.get("origem_acesso"),
        "avulsa_creditos":       int(u.get("avulsa_creditos", 0) or 0),
        "plano_solicitado":      plano_solicitado,
        "plano_solicitado_em":   u.get("plano_solicitado_em"),
        "plano_solicitado_nome": PLANOS.get(plano_solicitado, {}).get("nome"),
        "plano_ativo":           u.get("plano_ativo"),
        "usadas_hoje":           _contagem_hoje(u),
        "limite_hoje":           PLANOS.get(u.get("plano_ativo", ""), {}).get("importacoes_por_dia"),
        "pagamento_pendente": {
            "plano_id":  pendente.get("plano_id"),
            "url":       pendente.get("url"),
            "criado_em": pendente.get("criado_em"),
        } if pendente else None,
    })

@app.get("/assinatura/confirmar-pagamento")
async def assinatura_confirmar_pagamento(request: Request,
                                          order_nsu: str = "",
                                          transaction_nsu: str = "",
                                          slug: str = ""):
    sess = _sessao_ou_401(request)
    if not order_nsu:
        return err_json("order_nsu ausente.", 400)
    users = carregar_usuarios()
    _, u = _buscar_usuario(users, sess["usuario"])
    pendente = u.get("pagamento_pendente") if u else None
    if not pendente or pendente.get("order_nsu") != order_nsu:
        return ok_json({"ok": True, "pago": True, "msg": "Pagamento já confirmado."})
    ok, info = infinitepay_consultar_pagamento(order_nsu, transaction_nsu, slug)
    if not ok:
        return err_json(info)
    if not info.get("paid"):
        return ok_json({"ok": True, "pago": False, "msg": "Pagamento ainda não confirmado."})
    ok2, msg2 = processar_pagamento_confirmado(order_nsu, transaction_nsu, info.get("receipt_url", ""))
    return ok_json({"ok": ok2, "pago": ok2, "msg": msg2})

@app.get("/historico")
async def get_historico(request: Request):
    sess = _sessao_ou_401(request)
    historico = carregar_historico()
    historico_user = [h for h in historico if h.get("user_id") == sess["user_id"]]
    resumo = [{"nome": h["nome"], "total": h["total"], "salvo_em": h.get("salvo_em", "")}
              for h in historico_user]
    return ok_json({"ok": True, "historico": resumo})

@app.get("/historico/carregar")
async def historico_carregar(request: Request, nome: str = ""):
    sess = _sessao_ou_401(request)
    historico = carregar_historico()
    entrada = next((h for h in historico if h.get("nome") == nome and h.get("user_id") == sess["user_id"]), None)
    if entrada:
        return ok_json({"ok": True, **entrada})
    return err_json("Rota não encontrada no histórico.", 404)

@app.get("/coords/listar")
async def coords_listar(request: Request):
    _sessao_admin_ou_403(request)
    banco = banco_coords_carregar()
    entradas = [{"chave": k, **v} for k, v in sorted(banco["global"].items())]
    return ok_json({"ok": True, "total": len(entradas), "entradas": entradas})

@app.get("/admin/usuarios")
async def admin_usuarios(request: Request):
    _sessao_admin_ou_403(request)
    return ok_json({"ok": True, "usuarios": admin_listar_usuarios()})

@app.get("/admin/backup/status")
async def admin_backup_status(request: Request):
    _sessao_admin_ou_403(request)
    def _info(path_str):
        p = Path(path_str)
        recente = _snapshot_mais_recente(p)
        n_snapshots = len(list(BACKUP_DIR.glob(f"{p.stem}_*{p.suffix}")))
        return {
            "arquivo_existe": p.exists(),
            "tamanho_bytes": p.stat().st_size if p.exists() else 0,
            "modificado_em": datetime.fromtimestamp(p.stat().st_mtime).strftime("%d/%m/%Y %H:%M") if p.exists() else None,
            "snapshots_guardados": n_snapshots,
            "snapshot_mais_recente": recente.name if recente else None,
        }
    return ok_json({
        "ok": True,
        "usuarios": _info(USERS_FILE),
        "historico": _info(HISTORICO_FILE),
        "banco_coords": _info(BANCO_COORDS_FILE),
        "email_backup_configurado": bool(BACKUP_EMAIL_DESTINO),
        "email_destino": BACKUP_EMAIL_DESTINO or None,
        "intervalo_horas": BACKUP_EMAIL_INTERVALO_H,
    })

@app.post("/admin/backup/enviar-agora")
async def admin_backup_enviar_agora(request: Request):
    _sessao_admin_ou_403(request)
    corpo = await request.json() if request.headers.get("content-length", "0") != "0" else {}
    destino = (corpo or {}).get("email", "")
    ok, msg = enviar_backup_por_email(destino)
    if not ok:
        return err_json(msg, 400)
    return ok_json({"ok": True, "mensagem": msg})

@app.get("/admin/backup/baixar")
async def admin_backup_baixar(request: Request):
    _sessao_admin_ou_403(request)
    zip_path = _zip_backup_atual()
    return FileResponse(str(zip_path), media_type="application/zip", filename=zip_path.name)

# ═══════════════════════════════════════════════════════════════════
#  ROTAS  ──  POST
# ═══════════════════════════════════════════════════════════════════

@app.post("/auth/cadastro")
async def auth_cadastro(request: Request):
    data = await request.json()
    ok, msg, pending_token = iniciar_cadastro_pendente(
        data.get("usuario", ""), data.get("email", ""), data.get("senha", ""), data.get("telefone", "")
    )
    resp = {"ok": ok, "msg": msg}
    if ok:
        resp["pending_token"] = pending_token
    return ok_json(resp)

@app.post("/auth/confirmar-cadastro")
async def auth_confirmar_cadastro(request: Request):
    data = await request.json()
    ok, msg = confirmar_cadastro(data.get("pending_token", ""), data.get("codigo", ""))
    return ok_json({"ok": ok, "msg": msg})

@app.post("/auth/recuperar")
async def auth_recuperar(request: Request):
    data = await request.json()
    ok, msg, recovery_token = iniciar_recuperacao_senha(data.get("identificador", ""))
    if ok:
        return ok_json({"ok": True, "email_mascarado": msg, "recovery_token": recovery_token})
    return ok_json({"ok": False, "erro": msg})

@app.post("/auth/recuperar-confirmar")
async def auth_recuperar_confirmar(request: Request):
    data = await request.json()
    ok, msg = confirmar_codigo_recuperacao(data.get("recovery_token", ""), data.get("codigo", ""))
    return ok_json({"ok": ok, "msg": msg})

@app.post("/auth/recuperar-nova-senha")
async def auth_recuperar_nova_senha(request: Request):
    data = await request.json()
    ok, msg = redefinir_senha_recuperacao(data.get("recovery_token", ""), data.get("nova_senha", ""))
    return ok_json({"ok": ok, "msg": msg})

@app.post("/auth/login")
async def auth_login(request: Request):
    data = await request.json()
    usuario = data.get("usuario", "").strip()
    senha   = data.get("senha", "")
    resultado = autenticar_usuario(usuario, senha)
    if resultado:
        user_id, usuario_original = resultado
        is_admin = usuario_e_admin(usuario_original)
        token = criar_sessao(user_id, usuario_original, is_admin)
        return ok_json({"ok": True, "token": token, "usuario": usuario_original, "is_admin": is_admin})
    return ok_json({"ok": False, "erro": "Usuário ou senha incorretos."})

@app.post("/auth/logout")
async def auth_logout(request: Request):
    token = _token_da_request(request)
    if token:
        destruir_sessao(token)
    return ok_json({"ok": True})

@app.post("/auth/perfil/atualizar")
async def auth_perfil_atualizar(request: Request):
    sess = _sessao_ou_401(request)
    data = await request.json()
    telefone_raw  = data.get("telefone", "").strip()
    telefone_norm = _normalizar_telefone(telefone_raw)
    if telefone_norm and not _telefone_valido(telefone_norm):
        return err_json("Telefone inválido. Use DDD + 9 + número (ex: 62 9 91153473).")
    email_novo = data.get("email", "").strip()
    if email_novo and "@" not in email_novo:
        return err_json("E-mail inválido.")
    users = carregar_usuarios()
    chave, u = _buscar_usuario(users, sess["usuario"])
    if u is None:
        return err_json("Usuário não encontrado.")
    if telefone_norm:
        users[chave]["telefone"] = telefone_norm
    if email_novo:
        users[chave]["email"] = email_novo
    salvar_usuarios(users)
    return ok_json({"ok": True, "msg": "Perfil atualizado com sucesso."})

@app.post("/assinatura/solicitar")
async def assinatura_solicitar(request: Request):
    sess = _sessao_ou_401(request)
    data = await request.json()
    ok, msg = usuario_solicitar_plano(sess["usuario"], data.get("plano", ""))
    return ok_json({"ok": ok, "msg": msg})

@app.post("/assinatura/pagar")
async def assinatura_pagar(request: Request):
    sess = _sessao_ou_401(request)
    data = await request.json()
    ok, resultado = usuario_iniciar_pagamento(sess["usuario"], data.get("plano", ""), _base_url(request))
    if ok:
        return ok_json({"ok": True, "url": resultado})
    return err_json(resultado)

@app.post("/webhook/infinitepay")
async def webhook_infinitepay(request: Request):
    """Chamado pela InfinitePay servidor-a-servidor (sem token de sessão)."""
    try:
        payload = await request.json()
    except Exception:
        return JSONResponse({"success": False, "message": "JSON inválido."}, status_code=400)
    order_nsu = payload.get("order_nsu", "")
    if not order_nsu:
        return JSONResponse({"success": False, "message": "order_nsu ausente."}, status_code=400)
    ok, msg = processar_pagamento_confirmado(
        order_nsu,
        payload.get("transaction_nsu", ""),
        payload.get("receipt_url", ""),
    )
    if ok or msg == "Pedido não encontrado.":
        return JSONResponse({"success": True, "message": None})
    return JSONResponse({"success": False, "message": msg}, status_code=400)

@app.post("/admin/usuarios/criar")
async def admin_usuarios_criar(request: Request):
    sess = _sessao_admin_ou_403(request)
    data = await request.json()
    ok, msg = admin_criar_usuario(
        data.get("usuario", ""), data.get("senha", ""),
        data.get("email", ""), bool(data.get("is_admin", False))
    )
    return ok_json({"ok": ok, "msg": msg})

@app.post("/admin/usuarios/resetar-senha")
async def admin_resetar_senha_route(request: Request):
    _sessao_admin_ou_403(request)
    data = await request.json()
    ok, msg = admin_resetar_senha(data.get("usuario", ""), data.get("nova_senha", ""))
    return ok_json({"ok": ok, "msg": msg})

@app.post("/admin/usuarios/editar")
async def admin_editar_route(request: Request):
    _sessao_admin_ou_403(request)
    data = await request.json()
    ok, msg = admin_editar_contato(data.get("usuario", ""), data.get("email", ""), data.get("telefone", ""))
    return ok_json({"ok": ok, "msg": msg})

@app.post("/admin/usuarios/liberar-acesso")
async def admin_liberar_route(request: Request):
    _sessao_admin_ou_403(request)
    data = await request.json()
    ok, msg = admin_liberar_acesso(data.get("usuario", ""), data.get("dias", 0))
    return ok_json({"ok": ok, "msg": msg})

@app.post("/admin/usuarios/revogar-acesso")
async def admin_revogar_route(request: Request):
    _sessao_admin_ou_403(request)
    data = await request.json()
    ok, msg = admin_revogar_acesso(data.get("usuario", ""))
    return ok_json({"ok": ok, "msg": msg})

@app.post("/admin/usuarios/limpar-cache")
async def admin_limpar_cache_route(request: Request):
    _sessao_admin_ou_403(request)
    data = await request.json()
    ok, msg = admin_limpar_cache_usuario(data.get("usuario", ""))
    return ok_json({"ok": ok, "msg": msg})

@app.post("/admin/usuarios/limpar-cache-todos")
async def admin_limpar_cache_todos_route(request: Request):
    _sessao_admin_ou_403(request)
    ok, msg = admin_limpar_cache_todos_usuarios()
    return ok_json({"ok": ok, "msg": msg})

@app.post("/admin/usuarios/confirmar-plano")
async def admin_confirmar_plano_route(request: Request):
    _sessao_admin_ou_403(request)
    data = await request.json()
    ok, msg = admin_confirmar_plano(data.get("usuario", ""))
    return ok_json({"ok": ok, "msg": msg})

@app.post("/admin/usuarios/rejeitar-plano")
async def admin_rejeitar_plano_route(request: Request):
    _sessao_admin_ou_403(request)
    data = await request.json()
    ok, msg = admin_rejeitar_plano(data.get("usuario", ""))
    return ok_json({"ok": ok, "msg": msg})

@app.post("/coords/salvar")
async def coords_salvar(request: Request):
    sess = _sessao_ou_401(request)
    data = await request.json()
    endereco = data.get("endereco", "")
    ok, msg, info = banco_coords_salvar_coord(
        endereco, data.get("lat", 0), data.get("lon", 0),
        sess["user_id"]
    )
    resposta = {"ok": ok, "msg": msg}
    if ok:
        resposta.update(info)
        xp_resultado = _conceder_xp_endereco(sess["user_id"], endereco)
        if xp_resultado:
            resposta["gamificacao"] = xp_resultado
    else:
        resposta["erro"] = msg
    return ok_json(resposta)

@app.post("/coords/correlacionar")
async def coords_correlacionar(request: Request):
    """Chamado quando o usuário agrupa manualmente 2+ paradas (no mapa ou
    na lista) — salva que esses endereços (com texto diferente) são o
    mesmo lugar, como dica de consenso pra próxima vez."""
    sess = _sessao_ou_401(request)
    data = await request.json()
    enderecos = data.get("enderecos", [])
    if not isinstance(enderecos, list):
        return err_json("Campo 'enderecos' deve ser uma lista.")
    resultado = banco_coords_correlacionar(enderecos, sess["user_id"])
    return ok_json(resultado, 200 if resultado.get("ok") else 400)

@app.post("/coords/apagar")
async def coords_apagar(request: Request):
    _sessao_admin_ou_403(request)
    data = await request.json()
    ok, msg = banco_coords_apagar(data.get("endereco", ""))
    return ok_json({"ok": ok, "msg": msg})

@app.post("/coords/apagar-recentes")
async def coords_apagar_recentes(request: Request):
    """Zera do banco global tudo que foi confirmado nas últimas N horas
    (padrão 24h) — usado quando uma leva de geocodificações saiu errada."""
    _sessao_admin_ou_403(request)
    data = await request.json()
    try:
        horas = float(data.get("horas", 24))
    except (TypeError, ValueError):
        horas = 24
    total = banco_coords_apagar_recentes(horas)
    return ok_json({"ok": True, "removidas": total})

@app.post("/coords/apagar-intervalo")
async def coords_apagar_intervalo(request: Request):
    """Zera do banco global tudo confirmado num intervalo de datas
    específico (ex.: 28/06/2026 até hoje) — mesma ideia do apagar-recentes,
    só que com início e fim escolhidos em vez de só "últimas N horas"."""
    _sessao_admin_ou_403(request)
    data = await request.json()
    data_inicio = (data.get("data_inicio") or "").strip()
    data_fim = (data.get("data_fim") or "").strip()
    if not data_inicio or not data_fim:
        return err_json("Informe data_inicio e data_fim (YYYY-MM-DD).")
    try:
        total = banco_coords_apagar_intervalo(data_inicio, data_fim)
    except ValueError as e:
        return err_json(str(e))
    return ok_json({"ok": True, "removidas": total})

@app.post("/geocode/confirmar")
async def geocode_confirmar(request: Request):
    """
    Confirmação de geolocalização (tela "preparando sua rota"). Pra cada
    endereço recebido: 1) banco de coordenadas (override do usuário,
    senão o global já confirmado); 2) se não tiver, cai na cascata de
    CEP do csv_para_rota_xlsx.py (ViaCEP/BrasilAPI + Photon + Nominatim,
    com HERE por CEP estruturado só se a chave estiver configurada) —
    a mesma lógica do painel Anjun, aplicada aqui pra qualquer
    importação. Só o CEP entra na busca (nunca o endereço livre pro
    HERE/Google), porque nome de rua ambíguo é o que causava coordenada
    errada antes; achar pelo CEP dá o pino no lote/rua certa, no pior
    caso genérico mas nunca no bairro errado. Todo acerto vira entrada
    permanente no banco global.

    'enderecos' é um dict endereço_exibido -> texto_original: a rota Anjun
    tem o CEP removido do endereço já limpo (anjun_tratamento.py monta só
    rua+número, sem cidade/UF/CEP), então o CEP só sobrevive no texto
    original bruto — é dele, não do endereço de exibição, que o CEP é
    extraído aqui.
    """
    sess = _sessao_ou_401(request)
    usuario = sess.get("usuario", "")   # só pra registrar quem confirmou, no banco global
    user_id = sess["user_id"]           # chave do override pessoal (some quando a rota sai do histórico)
    data = await request.json()
    enderecos = data.get("enderecos") or {}
    if not isinstance(enderecos, dict):
        return err_json("Formato inválido: 'enderecos' deve ser um objeto endereço -> texto original.")
    itens = [
        (endereco.strip(), (texto_original or endereco).strip())
        for endereco, texto_original in enderecos.items()
        if isinstance(endereco, str) and endereco.strip()
    ][:300]  # limite de segurança por chamada

    resultados = {}
    pares_por_endereco = {}
    pares_unicos = set()

    for endereco, texto_original in itens:
        cache = banco_coords_buscar(endereco, user_id)
        if cache:
            resultados[endereco] = {"encontrado": True, **cache}
            continue
        cep = _extrair_cep(texto_original)
        if not cep:
            resultados[endereco] = {"encontrado": False}
            continue
        par = (cep, _extrair_numero(texto_original))
        pares_por_endereco[endereco] = par
        pares_unicos.add(par)

    if pares_unicos:
        coords_por_par = _geocodificar_enderecos(pares_unicos, "Goiânia", "GO")
        for endereco, par in pares_por_endereco.items():
            coord = coords_por_par.get(par)
            if not coord:
                resultados[endereco] = {"encontrado": False}
                continue
            lat, lon, fonte = coord
            banco_coords_salvar_global(endereco, lat, lon, fonte, usuario)
            resultados[endereco] = {"encontrado": True, "lat": lat, "lon": lon, "fonte": fonte}

    return ok_json({"ok": True, "resultados": resultados})

@app.post("/historico/atualizar-coords")
async def historico_atualizar_coords(request: Request):
    """Depois que o front confirma a geolocalização (HERE/Google) por cima
    dos dados que o /pipeline devolveu, manda as linhas atualizadas de
    volta aqui — pra refletir tanto no /dados desta sessão quanto no
    histórico já salvo. Sem isso, recarregar a rota (F5 ou pelo histórico)
    mostraria os badges como se nada tivesse sido confirmado."""
    sess = _sessao_ou_401(request)
    data = await request.json()
    rows = data.get("rows") or []
    if not isinstance(rows, list):
        return err_json("Formato inválido: 'rows' deve ser uma lista.")

    headers = sess["dados"][1] if sess.get("dados") else []
    sess["dados"] = (rows, headers)

    arq_final = ARQ_VALIDADO if Path(ARQ_VALIDADO).exists() else ARQ_PROCESSADO
    nome = Path(arq_final).name
    atualizado = atualizar_rows_historico(nome, rows, sess["user_id"])
    return ok_json({"ok": True, "historico_atualizado": atualizado})

@app.post("/upload")
async def upload(request: Request, arquivo: UploadFile = File(...)):
    """
    Recebe o arquivo xlsx via multipart/form-data.
    FastAPI + python-multipart fazem o parsing automaticamente —
    sem mais parsing manual de boundary.
    """
    sess = _sessao_com_acesso_ou_403(request)
    contents = await arquivo.read()
    if len(contents) <= 4:
        return err_json("Arquivo vazio ou inválido.")
    Path(ARQ_ENTRADA).write_bytes(contents)
    # guarda o hash do arquivo original pra dar XP uma única vez por rota (anti-fraude)
    sess["_rota_hash"] = gamification.calcular_hash_rota(contents)
    print(f"  [UPLOAD] {ARQ_ENTRADA} salvo ({len(contents)} bytes) — usuário: {sess['usuario']}")
    return ok_json({"ok": True})

@app.post("/scan/ocr")
async def scan_ocr(request: Request):
    """
    Recebe uma imagem em base64 (foto de etiqueta/pacote) e retorna o texto
    reconhecido via Google Cloud Vision API (OCR). Requer apenas login básico
    (não usa o gate de acesso pago do /upload, pois é uma feature separada).
    """
    _sessao_ou_401(request)

    if not GOOGLE_VISION_API_KEY:
        return err_json("OCR não configurado no servidor (GOOGLE_VISION_API_KEY ausente).", 500)

    data = await request.json()
    imagem_b64 = (data.get("imagem") or "").strip()
    if not imagem_b64:
        return err_json("Nenhuma imagem enviada.")

    # Remove prefixo data URL, se vier (ex.: "data:image/jpeg;base64,...")
    if "," in imagem_b64 and imagem_b64.strip().startswith("data:"):
        imagem_b64 = imagem_b64.split(",", 1)[1]

    payload = {
        "requests": [{
            "image": {"content": imagem_b64},
            "features": [{"type": "TEXT_DETECTION"}],
            "imageContext": {"languageHints": ["pt"]},
        }]
    }

    try:
        resp = requests.post(
            GOOGLE_VISION_URL,
            params={"key": GOOGLE_VISION_API_KEY},
            json=payload,
            timeout=30,
        )
        resp.raise_for_status()
        result = resp.json()
    except requests.exceptions.RequestException as e:
        print(f"  [SCAN/OCR] ❌ falha na chamada à Vision API: {e}")
        return err_json("Falha ao consultar o serviço de OCR. Tente novamente.", 502)

    resposta = (result.get("responses") or [{}])[0]
    if "error" in resposta:
        msg = resposta["error"].get("message", "Erro desconhecido na Vision API.")
        print(f"  [SCAN/OCR] ❌ Vision API: {msg}")
        return err_json(msg, 502)

    texto = ""
    annotations = resposta.get("textAnnotations") or []
    if annotations:
        texto = annotations[0].get("description", "")

    return ok_json({"ok": True, "texto": texto})

# ─── Anjun: CSV -> XLSX geocodificado, com progresso em tempo real ──
_ANJUN_JOBS: dict = {}  # uid -> {status, total, done, erro, pronto, nome_saida}

def _anjun_xlsx_para_csv(origem_path: Path, destino_path: Path):
    """
    Converte a aba ativa de um .xlsx pra .csv, célula por célula.
    Usado quando o usuário sobe o Anjun com xlsx em vez de csv — assim o
    csv_para_rota_xlsx.py continua recebendo sempre um CSV, sem precisar
    mexer nele.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl não instalado no servidor.")

    wb = load_workbook(origem_path, data_only=True)
    ws = wb.active
    with destino_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        linhas_gravadas = 0
        for row in ws.iter_rows(values_only=True):
            # Ignora linhas totalmente vazias (comuns no fim de planilhas exportadas)
            if row is None or all(v is None or str(v).strip() == "" for v in row):
                continue
            writer.writerow(["" if v is None else v for v in row])
            linhas_gravadas += 1
    if linhas_gravadas == 0:
        raise RuntimeError("a planilha está vazia.")

def _anjun_processar_bg(uid: str, entrada_path: Path, saida_path: Path):
    job = _ANJUN_JOBS[uid]
    try:
        env_here = {**os.environ, "HERE_API_KEY": HERE_API_KEY}
        proc = subprocess.Popen(
            [sys.executable, ANJUN_SCRIPT, str(entrada_path), str(saida_path)],
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, bufsize=1, env=env_here,
        )
        for linha in proc.stdout:
            linha = linha.rstrip("\n")
            if linha:
                print(f"  [ANJUN] {linha}")
            m_total = re.search(r"Geocodificando (\d+) CEP", linha)
            if m_total:
                job["total"] = int(m_total.group(1))
            m_item = re.match(r"\s*\[(\d+)/(\d+)\]", linha)
            if m_item:
                job["done"] = int(m_item.group(1))
                job["total"] = int(m_item.group(2))
        proc.wait(timeout=600)
        if proc.returncode != 0:
            job["erro"] = "Erro ao rodar o script (verifique o CSV de entrada)."
        elif not saida_path.exists():
            job["erro"] = "O script rodou mas não gerou o arquivo de saída."
        else:
            job["total"] = max(job["total"], job["done"], 1)
            job["done"] = job["total"]
            job["pronto"] = True
    except subprocess.TimeoutExpired:
        job["erro"] = "Timeout: a geocodificação demorou mais que o esperado."
    except Exception as e:
        job["erro"] = str(e)
    finally:
        job["status"] = "concluido"

@app.post("/anjun/iniciar")
async def anjun_iniciar(request: Request, arquivo: UploadFile = File(...)):
    """
    Recebe um CSV (formato delivery_list, sem sequência), salva e dispara o
    csv_para_rota_xlsx.py em segundo plano. Requer só login básico — é uma
    ferramenta separada do pipeline principal, não consome crédito de
    importação. O progresso é consultado via GET /anjun/progresso.
    """
    sess = _sessao_ou_401(request)

    if not Path(ANJUN_SCRIPT).exists():
        return err_json(f"{ANJUN_SCRIPT} não encontrado na pasta do servidor.")

    contents = await arquivo.read()
    if len(contents) <= 4:
        return err_json("Arquivo vazio ou inválido.")

    uid = sess["user_id"]
    sufixo_original = Path(arquivo.filename or "entrada.csv").suffix.lower() or ".csv"
    saida_path = DATA_DIR / f"anjun_saida_{uid}.xlsx"

    if sufixo_original == ".xls":
        return err_json(
            "Arquivo .xls (formato antigo do Excel) não é suportado. "
            "Salve como .xlsx ou .csv e envie de novo."
        )

    if sufixo_original == ".xlsx":
        # csv_para_rota_xlsx.py só sabe ler CSV — converte o xlsx enviado
        # antes de passar pro script, sem precisar alterá-lo.
        xlsx_tmp = DATA_DIR / f"anjun_entrada_{uid}_tmp.xlsx"
        entrada_path = DATA_DIR / f"anjun_entrada_{uid}.csv"
        xlsx_tmp.write_bytes(contents)
        try:
            _anjun_xlsx_para_csv(xlsx_tmp, entrada_path)
        except Exception as e:
            return err_json(f"Não consegui ler o xlsx enviado: {e}")
        finally:
            xlsx_tmp.unlink(missing_ok=True)
    else:
        entrada_path = DATA_DIR / f"anjun_entrada_{uid}{sufixo_original}"
        entrada_path.write_bytes(contents)

    nome_saida = f"{Path(arquivo.filename or 'anjun').stem}_GEO.xlsx"
    _ANJUN_JOBS[uid] = {
        "status": "rodando", "total": 0, "done": 0,
        "erro": None, "pronto": False, "nome_saida": nome_saida,
        "saida_path": str(saida_path),
    }

    print(f"  [ANJUN] {entrada_path.name} salvo ({len(contents)} bytes) — usuário: {sess['usuario']}")

    thread = threading.Thread(target=_anjun_processar_bg, args=(uid, entrada_path, saida_path), daemon=True)
    thread.start()

    return ok_json({"ok": True})

@app.get("/anjun/progresso")
async def anjun_progresso(request: Request):
    sess = _sessao_ou_401(request)
    job = _ANJUN_JOBS.get(sess["user_id"])
    if not job:
        return err_json("Nenhum processamento em andamento.")
    return ok_json({
        "ok": True,
        "status": job["status"],
        "total": job["total"],
        "done": job["done"],
        "pronto": job["pronto"],
        "erro": job["erro"],
    })

@app.get("/anjun/baixar")
async def anjun_baixar(request: Request):
    sess = _sessao_ou_401(request)
    job = _ANJUN_JOBS.get(sess["user_id"])
    if not job or not job.get("pronto"):
        return err_json("Arquivo ainda não está pronto.")
    saida_path = Path(job["saida_path"])
    if not saida_path.exists():
        return err_json("Arquivo de saída não encontrado (pode ter expirado).")
    return FileResponse(
        str(saida_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=job["nome_saida"],
    )

@app.post("/pipeline")
async def pipeline(request: Request):
    sess = _sessao_com_acesso_ou_403(request)
    if not Path(ARQ_ENTRADA).exists():
        return err_json(f"{ARQ_ENTRADA} não encontrado. Faça o upload primeiro.")
    if not Path(TRATAMENTO_PY).exists():
        return err_json(f"{TRATAMENTO_PY} não encontrado na pasta.")
    print(f"\n  [PIPELINE] Rodando {TRATAMENTO_PY}...")
    try:
        env_here = {**os.environ, "HERE_API_KEY": HERE_API_KEY, "HERE_CIDADE_UF": HERE_CIDADE_UF}
        result = subprocess.run(
            [sys.executable, TRATAMENTO_PY],
            capture_output=True, text=True, timeout=600, env=env_here
        )
        if result.returncode != 0:
            erro = result.stderr or result.stdout or "Erro desconhecido"
            print(f"  [PIPELINE] ❌ {erro}")
            return err_json(erro)
        print(f"  [PIPELINE] ✅ Pipeline concluído")
        if result.stdout:
            print(result.stdout)
        rows, headers = ler_processado(sess["user_id"])
        sess["dados"] = (rows, headers)
        arq_final = ARQ_VALIDADO if Path(ARQ_VALIDADO).exists() else ARQ_PROCESSADO
        adicionar_ao_historico(Path(arq_final).name, rows, headers, sess["user_id"])
        if not sess.get("is_admin"):
            usuario_consumir_credito_avulso_se_necessario(sess["usuario"])
            registrar_importacao_hoje(sess["usuario"])
        print(f"  [PIPELINE] ✅ {len(rows)} endereços carregados")

        xp_resultado = None
        rota_hash = sess.get("_rota_hash")
        if rota_hash:
            paradas = len(rows)
            pacotes = sum(int(r.get("group_size") or 1) for r in rows)
            xp_resultado = _conceder_xp_rota(sess["user_id"], paradas, pacotes, rota_hash)
            sess["_rota_hash"] = None  # cada upload só pode gerar XP uma vez

        resposta = {"ok": True, "total": len(rows)}
        if xp_resultado:
            resposta["gamificacao"] = xp_resultado
        return ok_json(resposta)
    except subprocess.TimeoutExpired:
        return err_json("Timeout: o pipeline demorou mais que o esperado.")
    except Exception as e:
        print(f"  [PIPELINE] ❌ {e}")
        return err_json(str(e))

# ─── Importar Rota Anjun: fluxo paralelo ao /upload + /pipeline ──────
# Recebe o CSV/xlsx bruto do Anjun, faz o tratamento Anjun (limpeza e
# deduplicação do endereço) e, em seguida, reaproveita a MESMA lógica de
# agrupamento/consolidação do tratamento_dados.py (anjun_tratamento.py
# importa tratamento_dados.py, não duplica nada dele) — depois já manda
# pra lista, sem precisar reimportar manualmente pelo botão principal.
# Não altera /upload, /pipeline nem tratamento_dados.py.
@app.post("/anjun-rota/importar")
async def anjun_rota_importar(request: Request, arquivo: UploadFile = File(...)):
    sess = _sessao_com_acesso_ou_403(request)

    if not Path(ANJUN_TRATAMENTO_PY).exists():
        return err_json(f"{ANJUN_TRATAMENTO_PY} não encontrado na pasta do servidor.")

    contents = await arquivo.read()
    if len(contents) <= 4:
        return err_json("Arquivo vazio ou inválido.")

    uid = sess["user_id"]
    sufixo_original = Path(arquivo.filename or "entrada.csv").suffix.lower() or ".csv"
    entrada_csv = DATA_DIR / f"anjunrota_entrada_{uid}.csv"
    saida_xlsx  = DATA_DIR / f"anjunrota_saida_{uid}.xlsx"

    if sufixo_original == ".xls":
        return err_json(
            "Arquivo .xls (formato antigo do Excel) não é suportado. "
            "Salve como .xlsx ou .csv e envie de novo."
        )

    if sufixo_original == ".xlsx":
        # anjun_tratamento.py só sabe ler CSV — converte o xlsx enviado
        # antes, reaproveitando o mesmo conversor do painel Anjun.
        xlsx_tmp = DATA_DIR / f"anjunrota_entrada_{uid}_tmp.xlsx"
        xlsx_tmp.write_bytes(contents)
        try:
            _anjun_xlsx_para_csv(xlsx_tmp, entrada_csv)
        except Exception as e:
            return err_json(f"Não consegui ler o xlsx enviado: {e}")
        finally:
            xlsx_tmp.unlink(missing_ok=True)
    else:
        entrada_csv.write_bytes(contents)

    print(f"  [ANJUN-ROTA] {entrada_csv.name} salvo ({len(contents)} bytes) — usuário: {sess['usuario']}")

    try:
        result = subprocess.run(
            [sys.executable, ANJUN_TRATAMENTO_PY, str(entrada_csv), str(saida_xlsx)],
            capture_output=True, text=True, timeout=180,
        )
        if result.returncode != 0:
            erro = result.stderr or result.stdout or "Erro desconhecido"
            print(f"  [ANJUN-ROTA] ❌ {erro}")
            return err_json(erro)
        if result.stdout:
            print(result.stdout)
        if not saida_xlsx.exists():
            return err_json("O tratamento rodou mas não gerou o arquivo de saída.")
    except subprocess.TimeoutExpired:
        return err_json("Timeout: o tratamento demorou mais que o esperado.")
    except Exception as e:
        print(f"  [ANJUN-ROTA] ❌ {e}")
        return err_json(str(e))

    # guarda o hash do arquivo original pra dar XP uma única vez por rota
    # (mesma lógica anti-fraude do /upload principal)
    sess["_rota_hash"] = gamification.calcular_hash_rota(contents)

    try:
        rows, headers = ler_processado(uid, caminho=saida_xlsx)
    except Exception as e:
        return err_json(f"Erro ao ler o resultado: {e}")

    sess["dados"] = (rows, headers)
    adicionar_ao_historico(saida_xlsx.name, rows, headers, uid)
    if not sess.get("is_admin"):
        usuario_consumir_credito_avulso_se_necessario(sess["usuario"])
        registrar_importacao_hoje(sess["usuario"])
    print(f"  [ANJUN-ROTA] ✅ {len(rows)} endereços carregados")

    xp_resultado = None
    rota_hash = sess.get("_rota_hash")
    if rota_hash:
        paradas = len(rows)
        pacotes = sum(int(r.get("group_size") or 1) for r in rows)
        xp_resultado = _conceder_xp_rota(uid, paradas, pacotes, rota_hash)
        sess["_rota_hash"] = None

    resposta = {"ok": True, "total": len(rows)}
    if xp_resultado:
        resposta["gamificacao"] = xp_resultado
    return ok_json(resposta)

# ─── Buscar Rota Anjun direto do site: sem upload manual ──────────────
# Roda anjun_extrator.py em segundo plano com o usuario/senha DO ANJUN que
# a pessoa digitou na hora (nunca os que estao no topo do .py) — assim cada
# usuario do Rota Manager usa a propria conta Anjun. anjun_extrator.py ja
# devolve o xlsx com endereco normalizado + lat/lng oficiais do Anjun,
# prontos pra cair direto em ler_processado(), igual ao /anjun-rota/importar.
_ANJUN_API_JOBS: dict = {}  # uid -> {status, total, done, erro, pronto, carregado, saida_path}

def _anjun_api_processar_bg(uid: str, usuario_anjun: str, senha_anjun: str, saida_path: Path):
    job = _ANJUN_API_JOBS[uid]
    try:
        env_anjun = {
            **os.environ,
            "PYTHONUNBUFFERED": "1",
            "ANJUN_USERNAME": usuario_anjun,
            "ANJUN_PASSWORD": senha_anjun,
            "ANJUN_OUTPUT_XLSX": str(saida_path),
            "ANJUN_OUTPUT_JSON": str(DATA_DIR / f"anjunapi_entrada_{uid}.json"),
            "ANJUN_OUTPUT_CSV": str(DATA_DIR / f"anjunapi_entrada_{uid}.csv"),
        }
        proc = subprocess.Popen(
            [sys.executable, ANJUN_EXTRATOR_PY],
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, bufsize=1, env=env_anjun,
        )
        for linha in proc.stdout:
            linha = linha.rstrip("\n")
            if linha:
                print(f"  [ANJUN-API] {linha}")
            m_total = re.match(r"\[ANJUN-API\] TOTAL (\d+)", linha)
            if m_total:
                job["total"] = int(m_total.group(1))
            m_item = re.match(r"\[ANJUN-API\] PROGRESSO (\d+)/(\d+)", linha)
            if m_item:
                job["done"] = int(m_item.group(1))
                job["total"] = int(m_item.group(2))
        proc.wait(timeout=600)
        if proc.returncode != 0:
            job["erro"] = "Não consegui buscar no Anjun (confira usuário e senha)."
        elif not saida_path.exists():
            job["erro"] = "Busca concluída, mas não há entregas pendentes no Anjun agora."
        else:
            job["total"] = max(job["total"], job["done"], 1)
            job["done"] = job["total"]
            job["pronto"] = True
    except subprocess.TimeoutExpired:
        job["erro"] = "Timeout: a busca no Anjun demorou mais que o esperado."
    except Exception as e:
        job["erro"] = str(e)
    finally:
        job["status"] = "concluido"

@app.post("/anjun-rota/buscar-api")
async def anjun_rota_buscar_api(request: Request):
    sess = _sessao_com_acesso_ou_403(request)

    if not Path(ANJUN_EXTRATOR_PY).exists():
        return err_json(f"{ANJUN_EXTRATOR_PY} não encontrado na pasta do servidor.")

    try:
        body = await request.json()
    except Exception:
        body = {}
    usuario_anjun = (body.get("usuario") or "").strip()
    senha_anjun = body.get("senha") or ""
    if not usuario_anjun or not senha_anjun:
        return err_json("Informe usuário e senha da sua conta Anjun.")

    uid = sess["user_id"]
    saida_path = DATA_DIR / f"anjunapi_saida_{uid}.xlsx"

    _ANJUN_API_JOBS[uid] = {
        "status": "rodando", "total": 0, "done": 0,
        "erro": None, "pronto": False, "carregado": False,
        "saida_path": str(saida_path),
    }

    print(f"  [ANJUN-API] busca iniciada — usuário Rota Manager: {sess['usuario']}")

    thread = threading.Thread(
        target=_anjun_api_processar_bg,
        args=(uid, usuario_anjun, senha_anjun, saida_path),
        daemon=True,
    )
    thread.start()

    return ok_json({"ok": True})

@app.get("/anjun-rota/api-progresso")
async def anjun_rota_api_progresso(request: Request):
    sess = _sessao_com_acesso_ou_403(request)
    uid = sess["user_id"]
    job = _ANJUN_API_JOBS.get(uid)
    if not job:
        return err_json("Nenhuma busca em andamento.")

    resposta = {
        "ok": True,
        "status": job["status"],
        "total": job["total"],
        "done": job["done"],
        "pronto": job["pronto"],
        "erro": job["erro"],
    }

    if job["pronto"] and not job.get("carregado"):
        saida_path = Path(job["saida_path"])
        try:
            rows, headers = ler_processado(uid, caminho=saida_path)
            sess["dados"] = (rows, headers)
            sess["_rota_hash"] = gamification.calcular_hash_rota(saida_path.read_bytes())
            adicionar_ao_historico(saida_path.name, rows, headers, uid)
            if not sess.get("is_admin"):
                usuario_consumir_credito_avulso_se_necessario(sess["usuario"])
                registrar_importacao_hoje(sess["usuario"])
            print(f"  [ANJUN-API] ✅ {len(rows)} endereços carregados — usuário: {sess['usuario']}")

            xp_resultado = None
            rota_hash = sess.get("_rota_hash")
            if rota_hash:
                paradas = len(rows)
                pacotes = sum(int(r.get("group_size") or 1) for r in rows)
                xp_resultado = _conceder_xp_rota(uid, paradas, pacotes, rota_hash)
                sess["_rota_hash"] = None
            if xp_resultado:
                resposta["gamificacao"] = xp_resultado

            resposta["total_enderecos"] = len(rows)
            job["carregado"] = True
        except Exception as e:
            job["erro"] = f"Erro ao carregar o resultado: {e}"
            resposta["erro"] = job["erro"]

    return ok_json(resposta)

# ─── Buscar Rota iMile direto do site: mesmo fluxo do Anjun acima ────────
# imile_extrator.py já devolve endereço + lat/lng oficiais do iMile (a API
# do app manda tudo pronto, sem precisar geocodificar), então cai direto em
# ler_processado() igual à Rota Anjun.
_IMILE_API_JOBS: dict = {}  # uid -> {status, total, done, erro, pronto, carregado, saida_path}

def _imile_api_processar_bg(uid: str, usuario_imile: str, senha_imile: str, saida_path: Path):
    job = _IMILE_API_JOBS[uid]
    try:
        env_imile = {
            **os.environ,
            "PYTHONUNBUFFERED": "1",
            "IMILE_USERNAME": usuario_imile,
            "IMILE_PASSWORD": senha_imile,
            "IMILE_OUTPUT_XLSX": str(saida_path),
            "IMILE_OUTPUT_JSON": str(DATA_DIR / f"imileapi_entrada_{uid}.json"),
            "IMILE_OUTPUT_CSV": str(DATA_DIR / f"imileapi_entrada_{uid}.csv"),
        }
        proc = subprocess.Popen(
            [sys.executable, IMILE_EXTRATOR_PY],
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, bufsize=1, env=env_imile,
        )
        for linha in proc.stdout:
            linha = linha.rstrip("\n")
            if linha:
                print(f"  [IMILE-API] {linha}")
            m_total = re.match(r"\[IMILE-API\] TOTAL (\d+)", linha)
            if m_total:
                job["total"] = int(m_total.group(1))
            m_item = re.match(r"\[IMILE-API\] PROGRESSO (\d+)", linha)
            if m_item:
                job["done"] = int(m_item.group(1))
                job["total"] = max(job["total"], job["done"])
        proc.wait(timeout=600)
        if proc.returncode != 0:
            job["erro"] = "Não consegui buscar no iMile (confira usuário e senha)."
        elif not saida_path.exists():
            job["erro"] = "Busca concluída, mas não há entregas pendentes no iMile agora."
        else:
            job["total"] = max(job["total"], job["done"], 1)
            job["done"] = job["total"]
            job["pronto"] = True
    except subprocess.TimeoutExpired:
        job["erro"] = "Timeout: a busca no iMile demorou mais que o esperado."
    except Exception as e:
        job["erro"] = str(e)
    finally:
        job["status"] = "concluido"

@app.post("/imile-rota/buscar-api")
async def imile_rota_buscar_api(request: Request):
    sess = _sessao_com_acesso_ou_403(request)

    if not Path(IMILE_EXTRATOR_PY).exists():
        return err_json(f"{IMILE_EXTRATOR_PY} não encontrado na pasta do servidor.")

    try:
        body = await request.json()
    except Exception:
        body = {}
    usuario_imile = (body.get("usuario") or "").strip()
    senha_imile = body.get("senha") or ""
    if not usuario_imile or not senha_imile:
        return err_json("Informe usuário e senha da sua conta iMile.")

    uid = sess["user_id"]
    saida_path = DATA_DIR / f"imileapi_saida_{uid}.xlsx"

    _IMILE_API_JOBS[uid] = {
        "status": "rodando", "total": 0, "done": 0,
        "erro": None, "pronto": False, "carregado": False,
        "saida_path": str(saida_path),
    }

    print(f"  [IMILE-API] busca iniciada — usuário Rota Manager: {sess['usuario']}")

    thread = threading.Thread(
        target=_imile_api_processar_bg,
        args=(uid, usuario_imile, senha_imile, saida_path),
        daemon=True,
    )
    thread.start()

    return ok_json({"ok": True})

@app.get("/imile-rota/api-progresso")
async def imile_rota_api_progresso(request: Request):
    sess = _sessao_com_acesso_ou_403(request)
    uid = sess["user_id"]
    job = _IMILE_API_JOBS.get(uid)
    if not job:
        return err_json("Nenhuma busca em andamento.")

    resposta = {
        "ok": True,
        "status": job["status"],
        "total": job["total"],
        "done": job["done"],
        "pronto": job["pronto"],
        "erro": job["erro"],
    }

    if job["pronto"] and not job.get("carregado"):
        saida_path = Path(job["saida_path"])
        try:
            rows, headers = ler_processado(uid, caminho=saida_path)
            sess["dados"] = (rows, headers)
            sess["_rota_hash"] = gamification.calcular_hash_rota(saida_path.read_bytes())
            adicionar_ao_historico(saida_path.name, rows, headers, uid)
            if not sess.get("is_admin"):
                usuario_consumir_credito_avulso_se_necessario(sess["usuario"])
                registrar_importacao_hoje(sess["usuario"])
            print(f"  [IMILE-API] ✅ {len(rows)} endereços carregados — usuário: {sess['usuario']}")

            xp_resultado = None
            rota_hash = sess.get("_rota_hash")
            if rota_hash:
                paradas = len(rows)
                pacotes = sum(int(r.get("group_size") or 1) for r in rows)
                xp_resultado = _conceder_xp_rota(uid, paradas, pacotes, rota_hash)
                sess["_rota_hash"] = None
            if xp_resultado:
                resposta["gamificacao"] = xp_resultado

            resposta["total_enderecos"] = len(rows)
            job["carregado"] = True
        except Exception as e:
            job["erro"] = f"Erro ao carregar o resultado: {e}"
            resposta["erro"] = job["erro"]

    return ok_json(resposta)

# ─── Importar Rota iMile (Circuit Detalhado): fluxo paralelo ao /upload ──
# Recebe o xlsx/csv "Circuit Detalhado" (LATITUDE, LONGITUDE, SEQUENCE +
# ENDERECO_ORIGINAL) exportado depois de rodar os endereços do Imile na
# Circuit, roda o tratamento Imile (desinverte número/rua, tira bairro/
# cidade/UF, já reaproveitando lat/long que vieram prontos) e cai na mesma
# lógica de agrupamento/consolidação do tratamento_dados.py — igual ao
# /anjun-rota/importar, mas pro export da Circuit em vez do Anjun.
@app.post("/imile-rota/importar")
async def imile_rota_importar(request: Request, arquivo: UploadFile = File(...)):
    sess = _sessao_com_acesso_ou_403(request)

    if not Path(IMILE_TRATAMENTO_PY).exists():
        return err_json(f"{IMILE_TRATAMENTO_PY} não encontrado na pasta do servidor.")

    contents = await arquivo.read()
    if len(contents) <= 4:
        return err_json("Arquivo vazio ou inválido.")

    uid = sess["user_id"]
    sufixo_original = Path(arquivo.filename or "entrada.xlsx").suffix.lower() or ".xlsx"
    if sufixo_original == ".xls":
        return err_json(
            "Arquivo .xls (formato antigo do Excel) não é suportado. "
            "Salve como .xlsx ou .csv e envie de novo."
        )

    # imile_tratamento.py lê .xlsx/.csv direto, sem precisar de conversão.
    entrada_path = DATA_DIR / f"imilerota_entrada_{uid}{sufixo_original}"
    saida_xlsx = DATA_DIR / f"imilerota_saida_{uid}.xlsx"
    entrada_path.write_bytes(contents)

    print(f"  [IMILE-ROTA] {entrada_path.name} salvo ({len(contents)} bytes) — usuário: {sess['usuario']}")

    try:
        result = subprocess.run(
            [sys.executable, IMILE_TRATAMENTO_PY, str(entrada_path), str(saida_xlsx)],
            capture_output=True, text=True, timeout=180,
        )
        if result.returncode != 0:
            erro = result.stderr or result.stdout or "Erro desconhecido"
            print(f"  [IMILE-ROTA] ❌ {erro}")
            return err_json(erro)
        if result.stdout:
            print(result.stdout)
        if not saida_xlsx.exists():
            return err_json("O tratamento rodou mas não gerou o arquivo de saída.")
    except subprocess.TimeoutExpired:
        return err_json("Timeout: o tratamento demorou mais que o esperado.")
    except Exception as e:
        print(f"  [IMILE-ROTA] ❌ {e}")
        return err_json(str(e))

    # guarda o hash do arquivo original pra dar XP uma única vez por rota
    # (mesma lógica anti-fraude do /upload principal)
    sess["_rota_hash"] = gamification.calcular_hash_rota(contents)

    try:
        rows, headers = ler_processado(uid, caminho=saida_xlsx)
    except Exception as e:
        return err_json(f"Erro ao ler o resultado: {e}")

    sess["dados"] = (rows, headers)
    adicionar_ao_historico(saida_xlsx.name, rows, headers, uid)
    if not sess.get("is_admin"):
        usuario_consumir_credito_avulso_se_necessario(sess["usuario"])
        registrar_importacao_hoje(sess["usuario"])
    print(f"  [IMILE-ROTA] ✅ {len(rows)} endereços carregados")

    xp_resultado = None
    rota_hash = sess.get("_rota_hash")
    if rota_hash:
        paradas = len(rows)
        pacotes = sum(int(r.get("group_size") or 1) for r in rows)
        xp_resultado = _conceder_xp_rota(uid, paradas, pacotes, rota_hash)
        sess["_rota_hash"] = None

    resposta = {"ok": True, "total": len(rows)}
    if xp_resultado:
        resposta["gamificacao"] = xp_resultado
    return ok_json(resposta)

@app.post("/rota/otimizar")
async def rota_otimizar(request: Request):
    """
    Reordena os endereços já carregados na sessão pela sequência mais
    eficiente de entrega, calculada via OSRM (dados OpenStreetMap).
    Endereços sem coordenada válida são mantidos, mas empurrados pro
    final da lista (não entram no cálculo de otimização).
    """
    sess = _sessao_ou_401(request)
    if sess["dados"] is None:
        return err_json("Nenhum dado processado ainda.", 404)
    rows, headers = sess["dados"]

    com_coord = []
    sem_coord = []
    for row in rows:
        try:
            lat = float(row.get("lat") or "")
            lon = float(row.get("lon") or "")
            com_coord.append((row, lat, lon))
        except (TypeError, ValueError):
            sem_coord.append(row)

    if len(com_coord) < 2:
        return err_json("É necessário pelo menos 2 endereços com coordenadas válidas para otimizar a rota.")

    coords = [(lat, lon) for _, lat, lon in com_coord]
    ordem, servico = _otimizar_sequencia(coords)
    if ordem is None:
        return err_json(
            "Não foi possível calcular a rota otimizada agora (serviços de roteamento indisponíveis). "
            "Tente novamente em instantes.", 502
        )

    rows_otimizadas = [com_coord[i][0] for i in ordem] + sem_coord
    sess["dados"] = (rows_otimizadas, headers)

    print(f"  [OTIMIZAR] {sess['usuario']} otimizou {len(com_coord)} paradas via {servico} "
          f"({len(sem_coord)} sem coordenada ficaram no final)")

    return ok_json({
        "ok": True,
        "rows": rows_otimizadas,
        "headers": headers,
        "sem_coordenadas": len(sem_coord),
        "servico": servico,
    })

# ═══════════════════════════════════════════════════════════════════
#  LOTES DE TERCEIROS (Aparecida de Goiânia / Senador Canedo / Goiânia)
#  ──────────────────────────────────────────────────────────────────
#  Proxy servidor-a-servidor para os vector tiles (.pbf) de quadra/lote
#  usados no Route Planner. Evita problema de CORS no navegador, esconde
#  o token do usuário final, e usa um cache SQLite PERMANENTE (em
#  DATA_DIR, mesmo Volume do Railway usado pro resto do app) — depois do
#  primeiro request (ou de rodar warmup_lotes.py), o tile nunca mais
#  precisa ser buscado de novo no routeplanner.com.br.
#  Lógica de cache/fetch fica em lotes_terceiros_cache.py, compartilhada
#  com o script de warm-up (warmup_lotes.py).
# ═══════════════════════════════════════════════════════════════════

from starlette.concurrency import run_in_threadpool
from lotes_terceiros_cache import get_tile as _lotes_get_tile
from lotes_terceiros_cache import LOTES_TERCEIROS_CIDADES as _LOTES_CIDADES

# ─── Rate limit dos tiles de quadra/lote ────────────────────────────
# Exigir login já derruba scraping anônimo; isso aqui é a segunda camada,
# pra tornar impraticável alguém com conta válida varrer um município
# inteiro de tile em tile. Janela deslizante em memória — mesmo padrão
# de _sessoes (single-instance; se algum dia rodar multi-instance no
# Railway, precisa virar Redis/DB compartilhado, senão cada instância
# conta separado).
_LOTES_RATE_LIMIT = 240   # tiles por janela — dá folga generosa pra navegação normal no mapa
_LOTES_RATE_WINDOW = 300  # segundos (5 min)
_lotes_rate_lock = threading.Lock()
_lotes_rate_hits: dict[str, list[float]] = {}

def _lotes_rate_check(usuario: str):
    agora = time.time()
    corte = agora - _LOTES_RATE_WINDOW
    with _lotes_rate_lock:
        hits = _lotes_rate_hits.setdefault(usuario, [])
        while hits and hits[0] < corte:
            hits.pop(0)
        if len(hits) >= _LOTES_RATE_LIMIT:
            raise HTTPException(status_code=429, detail="Muitas requisições de mapa em pouco tempo. Aguarde um pouco e tente de novo.")
        hits.append(agora)


@app.get("/api/lotes-terceiros/{cidade}/{z}/{x}/{y}.pbf")
async def lotes_terceiros_tile(cidade: str, z: int, x: int, y: int, request: Request):
    """Repassa um vector tile (.pbf) de quadra/lote de Aparecida de Goiânia,
    Senador Canedo ou Goiânia — lendo do cache permanente sempre que possível
    e só caindo pro Route Planner se o tile nunca foi visto antes.

    Exige sessão válida (evita scraping anônimo do dataset inteiro) e aplica
    rate limit por usuário (evita que uma conta válida varra um município
    inteiro de tile em tile em pouco tempo)."""
    sess = _sessao_ou_401(request)
    _lotes_rate_check(sess["usuario"])

    if cidade not in _LOTES_CIDADES:
        raise HTTPException(status_code=400, detail="Cidade inválida (use 'aparecida', 'canedo' ou 'goiania').")
    if not (13 <= z <= 17):
        raise HTTPException(status_code=400, detail="Zoom fora do intervalo permitido.")

    try:
        data = await run_in_threadpool(_lotes_get_tile, cidade, z, x, y)
    except Exception as e:
        print(f"  [LOTES-TERCEIROS] falha ao buscar tile {cidade} {z}/{x}/{y}: {type(e).__name__}: {e}")
        raise HTTPException(status_code=502, detail="Falha ao buscar tile.")

    if data is None:
        # Tile sem lote cadastrado nessa área — normal, não é erro (e já
        # ficou salvo no cache como tombstone, não busca de novo).
        return Response(status_code=204)

    return Response(
        content=data,
        media_type="application/x-protobuf",
        headers={"Cache-Control": "public, max-age=604800"},  # cache 7 dias no navegador
    )


# ─── Admin: indexação de quadra/lote (busca direta, sem precisar abrir mapa) ─
# Lê o MESMO lotes_cache.sqlite3 (populado pelo warmup_lotes.py) e monta a
# tabela `lotes_busca` (quadra + lote + centroide) pra busca instantânea.
# Reaproveita a autenticação admin já existente — nenhuma secret nova.
import io
from contextlib import redirect_stdout

try:
    import indexar_do_cache
    _INDEXAR_DISPONIVEL = True
except ImportError as _e:
    print(f"  [ADMIN/LOTES] ⚠️ indexar_do_cache indisponível ({_e}). "
          f"Adicione 'mapbox-vector-tile' e 'mercantile' ao requirements.txt.")
    indexar_do_cache = None
    _INDEXAR_DISPONIVEL = False


@app.post("/admin/lotes")
async def admin_lotes(request: Request):
    _sessao_admin_ou_403(request)

    if not _INDEXAR_DISPONIVEL:
        raise HTTPException(
            status_code=503,
            detail="indexar_do_cache indisponível: falta 'mapbox-vector-tile' e/ou "
                   "'mercantile' no requirements.txt. Adicione e faça o deploy de novo.",
        )

    body = await request.json()
    acao = body.get("acao")

    buffer = io.StringIO()
    try:
        if acao == "listar":
            with redirect_stdout(buffer):
                await run_in_threadpool(indexar_do_cache.listar)
        elif acao == "diagnostico":
            with redirect_stdout(buffer):
                await run_in_threadpool(indexar_do_cache.diagnostico)
        elif acao == "inspecionar":
            faltando = [k for k in ("cidade", "z", "x", "y") if body.get(k) is None]
            if faltando:
                raise HTTPException(status_code=400, detail=f"faltando: {', '.join(faltando)}")
            with redirect_stdout(buffer):
                await run_in_threadpool(
                    indexar_do_cache.inspecionar,
                    body["cidade"], int(body["z"]), int(body["x"]), int(body["y"]),
                )
        elif acao == "inspecionar_lote":
            faltando = [k for k in ("cidade", "quadra", "lote") if not body.get(k)]
            if faltando:
                raise HTTPException(status_code=400, detail=f"faltando: {', '.join(faltando)}")
            with redirect_stdout(buffer):
                await run_in_threadpool(
                    indexar_do_cache.inspecionar_lote,
                    body["cidade"], str(body["quadra"]), str(body["lote"]),
                    int(body.get("limite", 10)),
                )
        elif acao == "indexar":
            iniciou = indexar_do_cache.indexar_em_background()
            return JSONResponse({
                "ok": True,
                "iniciou": iniciou,
                "msg": "Indexação rodando em background. Consulte com acao=status."
                       if iniciou else "Já tem uma indexação rodando. Consulte com acao=status.",
            })
        elif acao == "status":
            return JSONResponse({"ok": True, **indexar_do_cache.obter_progresso()})
        elif acao == "consultar":
            # Consulta bruta em lotes_busca — não passa pelo parser/regex de
            # /buscar-lote, é só pra diagnosticar o que realmente está (ou
            # não está) gravado no banco.
            db_path = DATA_DIR / "lotes_cache.sqlite3"
            if not db_path.exists():
                raise HTTPException(status_code=503, detail="lotes_cache.sqlite3 não existe.")
            import sqlite3 as _sq_diag
            conn = _sq_diag.connect(str(db_path), timeout=10)
            try:
                if body.get("contagem_por_cidade"):
                    linhas = conn.execute(
                        "SELECT cidade, COUNT(*) FROM lotes_busca GROUP BY cidade"
                    ).fetchall()
                    return JSONResponse({"ok": True, "contagem_por_cidade": linhas})

                sql = "SELECT cidade, bairro, quadra, lote, via, busca_end, centroid_lat, centroid_lon FROM lotes_busca WHERE 1=1"
                params = []
                if body.get("cidade"):
                    sql += " AND cidade = ?"
                    params.append(body["cidade"])
                if body.get("quadra"):
                    sql += " AND quadra = ? COLLATE NOCASE"
                    params.append(str(body["quadra"]))
                if body.get("lote"):
                    sql += " AND lote = ? COLLATE NOCASE"
                    params.append(str(body["lote"]))
                if body.get("via_contem"):
                    sql += " AND via LIKE ?"
                    params.append(f"%{body['via_contem']}%")
                if body.get("bairro_contem"):
                    sql += " AND bairro LIKE ?"
                    params.append(f"%{body['bairro_contem']}%")
                sql += " LIMIT 50"
                linhas = conn.execute(sql, params).fetchall()
                return JSONResponse({"ok": True, "sql": sql, "resultados": linhas})
            finally:
                conn.close()
        else:
            raise HTTPException(status_code=400, detail="acao inválida: use listar | inspecionar | indexar | status | consultar")
    except HTTPException:
        raise
    except Exception as e:
        return JSONResponse(
            {"erro": f"{type(e).__name__}: {e}", "log": buffer.getvalue()}, status_code=500
        )

    return JSONResponse({"ok": True, "log": buffer.getvalue()})


# ─── Busca direta de quadra/lote (sem precisar procurar no mapa) ─────────────
# Consulta a tabela `lotes_busca` já pré-indexada (ver /admin/lotes acao=indexar).
# Aceita um único campo de texto livre (bom pro celular em campo) e tenta
# extrair quadra+lote de qualquer jeito que o usuário digitar.
import sqlite3 as _sqlite3_busca

# Padrão principal: "Q3 LT12", "Qd 270 lt 09", "Q 26, LT 43", "Q3-LT12"...
# Captura só dígitos (+ sufixo de 1 letra opcional, ex: "3A") pra não deixar
# o grupo da quadra "vazar" pra dentro do "LT" quando não tem espaço/vírgula
# separando (bug antigo: "Q3 LT12" lia lote="2" em vez de "12").
_RE_TERMO_Q_LT = re.compile(
    r"Q[A-Za-z]*\.?\s*([0-9]+[A-Za-z]?)\s*[-,/]?\s*L[A-Za-z]*\.?\s*([0-9]+[A-Za-z]?)",
    re.IGNORECASE,
)
# Formato "270-9" (o mesmo usado no busca_end formatado, ex:
# "AVENIDA C104, 270-9, JARDIM AMÉRICA") — sem prefixo Q/LT.
_RE_TERMO_HIFEN = re.compile(r"\b(\d+)\s*-\s*(\d+)\b")
_RE_TERMO_DOIS_NUMEROS = re.compile(r"(\d+[A-Za-z]?)\D+?(\d+[A-Za-z]?)")


def _normalizar_num_busca(v):
    """Tira zero à esquerda ('09' -> '9') pra bater com o jeito que o
    índice grava e o usuário digita. Mantém sufixo de letra em maiúsculo."""
    if v is None:
        return None
    v = str(v).strip()
    if not v:
        return None
    return str(int(v)) if v.isdigit() else v.upper()


def _parse_termo_busca_lote(termo: str):
    """Tenta extrair (quadra, lote, resto, resto_filtro) de um texto livre
    digitado pelo usuário. Aceita formatos como 'Q3 LT12', 'quadra 3 lote
    12', '3/12', '270-9', 'jardim tropical q3, lt12'. Retorna
    (None, None, termo, termo) se não achar quadra/lote."""
    termo = (termo or "").strip()
    if not termo:
        return None, None, termo, termo

    m = _RE_TERMO_Q_LT.search(termo)
    if not m:
        m = _RE_TERMO_HIFEN.search(termo)
    if not m:
        m = _RE_TERMO_DOIS_NUMEROS.search(termo)
    if not m:
        return None, None, termo, termo

    quadra, lote = m.group(1), m.group(2)
    resto = (termo[:m.start()] + " " + termo[m.end():]).strip()
    # Se sobrou algo tipo "AVENIDA C104,  , JARDIM AMÉRICA", o pedaço que
    # mais importa pro filtro de bairro é o último depois da vírgula.
    partes = [p.strip() for p in resto.split(",") if p.strip()]
    resto_filtro = partes[-1] if partes else resto
    return _normalizar_num_busca(quadra), _normalizar_num_busca(lote), resto, resto_filtro


@app.get("/buscar-lote")
async def buscar_lote(request: Request, q: str = "", cidade: str = "",
                       perto_lat: float | None = None, perto_lon: float | None = None):
    _sessao_ou_401(request)

    quadra, lote, resto, resto_filtro = _parse_termo_busca_lote(q)
    if not quadra or not lote:
        raise HTTPException(
            status_code=400,
            detail="Não consegui identificar quadra e lote no termo. "
                   "Tente algo como 'Q3 LT12', '3/12' ou 'jardim tropical q3 lt12'.",
        )

    db_path = DATA_DIR / "lotes_cache.sqlite3"
    if not db_path.exists():
        raise HTTPException(
            status_code=503,
            detail="Ainda não existe índice de quadra/lote. Rode a indexação em "
                   "/admin (acao=indexar) primeiro.",
        )

    try:
        conn = _sqlite3_busca.connect(str(db_path), timeout=10)
        base_sql = (
            "SELECT cidade, bairro, quadra, lote, via, busca_end, "
            "centroid_lat, centroid_lon FROM lotes_busca "
            "WHERE quadra = ? COLLATE NOCASE AND lote = ? COLLATE NOCASE"
        )
        base_params = [quadra, lote]
        if cidade:
            base_sql += " AND cidade = ?"
            base_params.append(cidade)

        linhas = []
        if resto_filtro:
            # Primeiro tenta refinar por bairro/via. Só usa esse resultado
            # se achou algo — senão cai pro fallback sem filtro, porque
            # cidade sem bairro cadastrado no tile (ex: Goiânia) nunca
            # bateria no LIKE e o lote sumiria da busca sem motivo.
            sql_refinado = base_sql + " AND (bairro LIKE ? OR via LIKE ?) ORDER BY cidade, bairro LIMIT 50"
            params_refinado = base_params + [f"%{resto_filtro}%", f"%{resto_filtro}%"]
            linhas = conn.execute(sql_refinado, params_refinado).fetchall()

        if not linhas:
            sql_fallback = base_sql + " ORDER BY cidade, bairro LIMIT 50"
            linhas = conn.execute(sql_fallback, base_params).fetchall()

        conn.close()
    except _sqlite3_busca.OperationalError as e:
        raise HTTPException(
            status_code=503,
            detail=f"Índice de lotes indisponível ou desatualizado: {e}. "
                   f"Rode a indexação em /admin (acao=indexar).",
        )

    # Quando a mesma quadra/lote existe em mais de uma cidade (ex: "Q212
    # LT1" existe tanto em Goiânia quanto em Aparecida), a ordem alfabética
    # de cidade sozinha não diz qual é a certa. Desempata contando quantas
    # palavras do texto digitado (rua, bairro etc) aparecem no bairro/via
    # de cada linha, e manda o melhor match pro topo da lista.
    if len(linhas) > 1 and resto:
        _tokens = [t for t in re.split(r"[^0-9A-Za-zÀ-ÿ]+", resto.lower()) if len(t) > 2]

        def _pontuacao(linha):
            texto = f"{linha[1] or ''} {linha[4] or ''}".lower()
            return sum(1 for t in _tokens if t in texto)
    else:
        def _pontuacao(linha):
            return 0

    # Segundo critério: distância até a coordenada atual do pin (quando
    # informada). Cobre o caso de cidades sem bairro/via cadastrado no
    # tile (ex: Goiânia) — ali a pontuação por palavra empata em zero pra
    # todo mundo, e a proximidade com o pin já colocado é o melhor sinal
    # que temos pra escolher entre lotes de mesma quadra/lote espalhados
    # pela cidade.
    def _distancia(linha):
        if perto_lat is None or perto_lon is None or linha[6] is None or linha[7] is None:
            return float("inf")
        return ((linha[6] - perto_lat) ** 2 + (linha[7] - perto_lon) ** 2) ** 0.5

    if len(linhas) > 1:
        linhas = sorted(linhas, key=lambda l: (-_pontuacao(l), _distancia(l), l[0], l[1] or ""))

    def _distancia_metros(lat_r, lon_r):
        if perto_lat is None or perto_lon is None or lat_r is None or lon_r is None:
            return None
        import math
        R = 6371000  # raio da Terra em metros
        p1, p2 = math.radians(perto_lat), math.radians(lat_r)
        dphi = math.radians(lat_r - perto_lat)
        dlmb = math.radians(lon_r - perto_lon)
        a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlmb / 2) ** 2
        return round(R * 2 * math.atan2(a ** 0.5, (1 - a) ** 0.5))

    resultados = [
        {
            "cidade": r[0], "bairro": r[1], "quadra": r[2], "lote": r[3],
            "via": r[4], "busca_end": r[5], "lat": r[6], "lon": r[7],
            "distancia_m": _distancia_metros(r[6], r[7]),
        }
        for r in linhas if r[6] is not None and r[7] is not None
    ]
    return ok_json({"ok": True, "quadra": quadra, "lote": lote, "resultados": resultados})


# ═══════════════════════════════════════════════════════════════════
#  ROTAS  ──  DELETE
# ═══════════════════════════════════════════════════════════════════

@app.delete("/admin/usuarios")
async def admin_apagar_usuario_route(request: Request, usuario: str = ""):
    sess = _sessao_admin_ou_403(request)
    ok, msg = admin_apagar_usuario(usuario, sess["usuario"])
    return ok_json({"ok": ok, "msg": msg})

@app.delete("/historico/apagar")
async def historico_apagar(request: Request, nome: str = ""):
    sess = _sessao_ou_401(request)
    historico = carregar_historico()
    existia = any(h.get("nome") == nome and h.get("user_id") == sess["user_id"] for h in historico)
    novo = [h for h in historico if not (h.get("nome") == nome and h.get("user_id") == sess["user_id"])]
    salvar_historico(novo)
    if existia:
        # Rota saiu do histórico dele — as correções manuais que ele fez
        # nela eram só temporárias, então voltam ao valor global normal.
        banco_coords_limpar_overrides_usuario(sess["user_id"])
    return ok_json({"ok": True})

# ═══════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════

# Registrado por último de propósito — ver comentário lá em cima, perto
# do app.mount de /static/icons.
app.include_router(gamification.router)

if __name__ == "__main__":
    print(f"""
╔══════════════════════════════════════════════════╗
║       ROTA MANAGER — SERVIDOR  (FastAPI)         ║
╠══════════════════════════════════════════════════╣
║  Endereço : http://{HOST}:{PORT}
║  Pasta    : {Path('.').resolve()}
╚══════════════════════════════════════════════════╝
""")
    try:
        import openpyxl  # noqa
    except ImportError:
        print("⚠️  openpyxl não encontrado. Instalando...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])

    _migrar_para_volume()
    _bootstrap_admin()

    threading.Thread(target=_loop_backup_periodico, daemon=True).start()

    uvicorn.run(
        app,
        host=HOST,
        port=PORT,
        workers=1,          # múltiplos workers incompatíveis com sessões em memória
        log_level="info",
    )
