"""
imile_tratamento.py
====================
Fluxo PARALELO ao pipeline principal, feito sob medida pro export
"Circuit Detalhado" que sai depois de rodar os endereços do Imile na
Circuit (ex.: rota_CIRCUIT_DETALHADO.xlsx): 3 colunas — LATITUDE,
LONGITUDE e uma coluna única "SEQUENCE + ENDERECO_ORIGINAL" no formato

    37 - SN,Rua 2 Avenida Antônio Martins Borges quadra 113 lote 09,Setor Pedro Ludovico,Goiânia,Goiás
    43 - 79,Avenida Edmundo Pinheiro de Abreu Ap 102 edifício Santiago,Setor Pedro Ludovico,Goiânia,Goiás

O problema que este script resolve
-----------------------------------
Diferente do Anjun, o Imile/Circuit costuma inverter o NÚMERO (ou "SN")
pra ANTES do nome da rua, separado só por vírgula — ex.: "79,Avenida
Edmundo..." em vez de "Avenida Edmundo..., 79". Passado direto pro
`standardize()` do tratamento_dados.py, esse número vira parte do "nome
da rua" e o endereço sai bagunçado. Às vezes também aparece uma via de
referência solta ANTES do nome de rua real (ex.: "Rua 2 Avenida Antônio
Martins Borges..." — "Rua 2" é só uma travessa citada de passagem).

Diferente do Anjun, aqui a gente JÁ TEM lat/long por linha (vieram do
geocoding da própria Circuit) — não precisamos geocodificar de novo, só
carregamos coord/lat/lon junto com cada linha pra consolidação. Quando a
Circuit falhou (lat/long = 0,0), a linha fica sem coordenada e entra no
contador "sem coordenada" no resumo final, pra você saber quantas ainda
precisam de geocoding manual/fallback.

Passo 1 (IMILE) — feito aqui: separa SEQUENCE do endereço, tira bairro/
cidade/UF do fim (sempre os 3 últimos campos separados por vírgula,
mesmo quando o "bairro" veio bugado repetindo o nome do estado),
desinverte número-antes-da-rua e remove via de referência solta antes do
nome real. Devolve endereço no formato "rua, número" / "rua, QD LT" que
o tratamento_dados.py sabe ler — reaproveitando `normalizar_detalhe_anjun`
do anjun_tratamento.py pra extração de número/QD-LT/nome de rua (mesma
lógica, só muda o jeito de chegar no texto "endereço puro").

Passo 2 (ATUAL) — reaproveita `group_rows_p1` + `consolidate_p2` +
`write_excel_p2` do tratamento_dados.py, sem alterar nada nele.

Uso:
    python imile_tratamento.py entrada.xlsx saida.xlsx

Entrada esperada: .xlsx (ou .csv) com colunas LATITUDE, LONGITUDE e uma
coluna contendo "SEQUENCE" no nome, no formato
"N - endereco,Bairro,Cidade,UF".

Limitação conhecida: a remoção de "via de referência solta antes da rua
real" só trata UM prefixo cruzado (ex.: "Rua 2 Avenida X..."). Se
aparecerem casos com dois prefixos empilhados, ou outros padrões novos
do Imile, me manda um print/linha de exemplo que eu ajusto o regex.
"""

import csv
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

import tratamento_dados as td
import anjun_tratamento as anj

try:
    import gemini_enderecos
except ImportError:
    gemini_enderecos = None


_RE_SEQ_ENDERECO = re.compile(r'^\s*(\d+)\s*-\s*(.*)$', re.DOTALL)

# Número (ou "SN") ANTES do nome da rua, separados só por vírgula —
# padrão de inversão típico do Imile: "79,Avenida X" / "SN,Rua Y".
_RE_INVERTIDO = re.compile(r'^(?:(\d{1,6})|SN)\s*,\s*(.+)$', re.IGNORECASE | re.DOTALL)

# Rua-código SEM nenhum prefixo (nem "Rua", nem "R.") — comum no Setor
# Pedro Ludovico, onde as ruas já são só números: "1024  434 Ap 101" em
# vez de "Rua 1024, 434 Ap 101". Sem isso, o primeiro número (código da
# rua) e o segundo (número da casa) ficam colados e o parser do
# tratamento_dados.py não consegue separar um do outro.
_RE_CODIGO_SEM_PREFIXO = re.compile(r'^(\d{3,5})\s+(\d{1,5}\b.*)$', re.DOTALL)


def _remover_prefixo_cruzado(texto: str) -> str:
    """Às vezes o Imile grava uma via de referência ANTES do endereço
    real — ex.: 'Rua 2 Avenida Antônio Martins Borges quadra 113 lote
    09': 'Rua 2' é só uma travessa citada de passagem, o endereço de
    verdade é a 'Avenida Antônio Martins Borges'. Sinal disso: logo
    depois do primeiro prefixo+código (que bateria em _STREET_HEAD_RE)
    vem OUTRO prefixo de rua reconhecível — nesse caso descartamos o
    primeiro trecho e ficamos só com o segundo."""
    m1 = anj._STREET_HEAD_RE.match(texto)
    if not m1:
        return texto
    resto = texto[m1.end():].strip().lstrip(',').strip()
    if anj._PREFIXO_RUA_RE.match(resto):
        return resto
    return texto


def _dividir_endereco_bairro_cidade_uf(texto: str):
    """O texto após 'N - ' sempre termina em '...,Bairro,Cidade,UF' (3
    campos fixos no final, mesmo quando o próprio endereço tem vírgulas
    no meio por causa da inversão número/rua). Quando o "bairro" real
    não veio (o Imile às vezes repete o nome da cidade/estado no lugar),
    tratamos como bairro vazio pra não sujar o agrupamento por bairro do
    tratamento_dados.py."""
    partes = [p.strip() for p in texto.split(',')]
    if len(partes) >= 4:
        uf, cidade, bairro = partes[-1], partes[-2], partes[-3]
        endereco = ','.join(partes[:-3]).strip()
    elif len(partes) == 3:
        uf, cidade, bairro = partes[-1], partes[-2], ''
        endereco = partes[0].strip()
    else:
        endereco, bairro, cidade, uf = texto.strip(), '', '', ''

    if bairro and td.strip_accents(bairro).strip().lower() in (
        td.strip_accents(cidade).strip().lower(),
        td.strip_accents(uf).strip().lower(),
    ):
        bairro = ''

    return endereco, bairro, cidade, uf


def extrair_endereco_imile(endereco_bruto: str) -> str:
    """Recebe o trecho de endereço (sem seq/bairro/cidade/UF) e devolve
    uma versão limpa, já desinvertida, pronta pro `standardize()` do
    tratamento_dados.py."""
    if not endereco_bruto or not endereco_bruto.strip():
        return endereco_bruto or ''
    texto = endereco_bruto.strip()
    # "214. s/n 204" - ponto solto depois de numero (erro de digitacao
    # comum, tipo "214." em vez de "214,") atrapalha a extracao do
    # numero mais adiante; remove antes de qualquer outra coisa.
    texto = re.sub(r'(?<=\d)\.(?=\s)', '', texto)

    numero_prefixo = None
    m_inv = _RE_INVERTIDO.match(texto)
    if m_inv:
        candidato_num, resto = m_inv.group(1), m_inv.group(2).strip()
        if anj._PREFIXO_RUA_RE.match(resto) or anj._STREET_HEAD_RE.match(resto):
            numero_prefixo = candidato_num  # None quando era "SN,"
            texto = resto
    texto = _remover_prefixo_cruzado(texto)

    # Rua-codigo sem NENHUM prefixo reconhecivel (nem "Rua", nem sigla) -
    # ex.: "1024  434 Ap 101". Reconstroi como "RUA 1024, 434 Ap 101"
    # pra cair no fluxo normal de rua-codificada do anjun_tratamento.
    if not (anj._PREFIXO_RUA_RE.match(texto) or anj._STREET_HEAD_RE.match(texto)):
        m_cod = _RE_CODIGO_SEM_PREFIXO.match(texto)
        if m_cod:
            texto = f"RUA {m_cod.group(1)}, {m_cod.group(2)}"

    return anj.normalizar_detalhe_anjun(texto, numero_prefixo)


# ── Leitura do arquivo Circuit Detalhado ──

def _achar_coluna_sequence(headers: list) -> int:
    for i, h in enumerate(headers):
        if 'sequence' in (h or '').lower():
            return i
    raise ValueError("Coluna 'SEQUENCE + ENDERECO_ORIGINAL' não encontrada.")


def _ler_linhas(caminho: Path) -> list:
    """Lê .xlsx (Circuit Detalhado) ou .csv com colunas LATITUDE,
    LONGITUDE e SEQUENCE+ENDERECO. Devolve lista de dicts lat/lon/raw."""
    if caminho.suffix.lower() in ('.xlsx', '.xlsm'):
        wb = load_workbook(caminho, data_only=True)
        ws = wb.active
        headers = [str(c.value or '').strip() for c in ws[1]]
        idx_lat = next((i for i, h in enumerate(headers) if 'lat' in h.lower()), None)
        idx_lon = next((i for i, h in enumerate(headers) if 'lon' in h.lower()), None)
        idx_seq = _achar_coluna_sequence(headers)

        linhas = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            def g(i):
                if i is None or i >= len(row) or row[i] is None:
                    return ''
                return str(row[i]).strip()
            linhas.append({'lat': g(idx_lat), 'lon': g(idx_lon), 'raw': g(idx_seq)})
        return linhas

    with open(caminho, encoding='utf-8-sig', newline='') as f:
        leitor = csv.DictReader(f)
        headers = leitor.fieldnames or []
        col_seq = next((h for h in headers if 'sequence' in h.lower()), None)
        col_lat = next((h for h in headers if 'lat' in h.lower()), None)
        col_lon = next((h for h in headers if 'lon' in h.lower()), None)
        if not col_seq:
            raise ValueError("Coluna 'SEQUENCE + ENDERECO_ORIGINAL' não encontrada no CSV.")
        return [
            {
                'lat': (linha.get(col_lat) or '').strip() if col_lat else '',
                'lon': (linha.get(col_lon) or '').strip() if col_lon else '',
                'raw': (linha.get(col_seq) or '').strip(),
            }
            for linha in leitor
        ]


def _linhas_para_passo1(linhas: list) -> list:
    rows = []
    for linha in linhas:
        m_seq = _RE_SEQ_ENDERECO.match(linha['raw'])
        if not m_seq:
            continue
        seq, resto = m_seq.group(1), m_seq.group(2).strip()

        endereco_bruto, bairro, cidade, uf = _dividir_endereco_bairro_cidade_uf(resto)
        endereco_limpo = extrair_endereco_imile(endereco_bruto)

        lat, lon = linha['lat'], linha['lon']
        # (0, 0) é falha de geocoding da Circuit — não é coordenada válida
        if lat in ('0', '0.0') and lon in ('0', '0.0'):
            lat, lon = '', ''
        coord = f"{lat},{lon}" if lat and lon else ''

        rows.append({
            'raw_row': [],
            'stop': seq,
            'address': endereco_limpo,
            'coord': coord,
            '_original': endereco_bruto,
            '_bairro': bairro,
            '_cidade': cidade,
            '_uf': uf,
            '_lat': lat,
            '_lon': lon,
        })
    return rows


def _passo1_para_passo2(rows_agrupados: list) -> list:
    saida = []
    for r in rows_agrupados:
        saida.append({
            'reformado': r.get('group_label', '') or r.get('address', ''),
            'original': r.get('_original', ''),
            'bairro': r.get('_bairro', ''),
            'seq': r.get('stop', ''),
            'stop': r.get('stop', ''),
            'coord': r.get('coord', ''),
            'zip': '',
            'lat': r.get('_lat', ''),
            'lon': r.get('_lon', ''),
            'contato': '',
        })
    return saida


def processar_imile(caminho_entrada: str, caminho_saida: str, usar_gemini: bool = True) -> str:
    caminho_entrada = Path(caminho_entrada)
    if not caminho_entrada.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho_entrada}")

    linhas = _ler_linhas(caminho_entrada)
    if not linhas:
        raise ValueError("Arquivo vazio ou sem linhas reconhecíveis.")

    print(f"Lendo: {caminho_entrada} ({len(linhas)} linha(s))")

    print("Fazendo o tratamento Imile (desinvertendo/limpando endereços)...")
    rows_p1 = _linhas_para_passo1(linhas)

    print("Agrupando (mesma lógica do tratamento_dados.py)...")
    rows_p1 = td.group_rows_p1(rows_p1)

    print("Consolidando...")
    rows_p2 = _passo1_para_passo2(rows_p1)
    groups = td.consolidate_p2(rows_p2)

    if usar_gemini and gemini_enderecos is not None:
        print("Corrigindo enderecos com Gemini...")
        brutos = [(g['original'].split(' | ', 1)[0] if g['original'] else g['key']) for g in groups]
        corrigidos = gemini_enderecos.limpar_enderecos(brutos)
        for g, corrigido in zip(groups, corrigidos):
            if corrigido:
                g['key'] = corrigido

    caminho_saida = Path(caminho_saida)
    print(f"Salvando: {caminho_saida}")
    td.write_excel_p2(groups, str(caminho_saida))

    sem_coord = sum(1 for r in rows_p1 if not r.get('coord'))
    print(f"✅ {len(linhas)} linha(s) -> {len(groups)} endereço(s) único(s) "
          f"| {sum(1 for g in groups if g['count'] > 1)} com duplicatas"
          f"{f' | {sem_coord} sem coordenada (falha Circuit)' if sem_coord else ''}")

    return str(caminho_saida)


def main():
    if len(sys.argv) < 3:
        print("Uso: python imile_tratamento.py entrada.xlsx saida.xlsx")
        sys.exit(1)
    processar_imile(sys.argv[1], sys.argv[2])


if __name__ == '__main__':
    main()
